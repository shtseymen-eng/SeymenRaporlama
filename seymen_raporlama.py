# ╔══════════════════════════════════════════════════════════════════════════╗
# ║          SEYMEN RAPORLAMA  —  Pregate Operasyon Yönetim Sistemi         ║
# ║                        Sürüm 1.0  |  S. SEYMEN                         ║
# ╚══════════════════════════════════════════════════════════════════════════╝
#
#  KURULUM TALİMATLARI:
#  ─────────────────────────────────────────────────────────────────────────
#  1) Python 3.9 veya üstü kurulu olmalı  →  https://www.python.org
#
#  2) Terminali/Komut İstemini açın ve şunu yapıştırın:
#       pip install customtkinter pandas openpyxl python-pptx pillow
#
#  3) Programı çalıştırın:
#       python seymen_raporlama.py
#
#  MASAÜSTÜ KISAYOLU (Windows):
#  ─────────────────────────────────────────────────────────────────────────
#  → seymen_raporlama.py dosyasına sağ tıklayın
#  → "Gönder"  →  "Masaüstü (kısayol oluştur)" seçin
#  → Masaüstündeki kısayola çift tıklayın → program açılır
#
#  MASAÜSTÜ KISAYOLU (macOS):
#  ─────────────────────────────────────────────────────────────────────────
#  → seymen_raporlama.py dosyasını Finder'da bulun
#  → Dosyayı masaüstüne sürükleyin  (veya Alias oluşturun)
#  → Çift tıklayarak açın
#
#  MASAÜSTÜ KISAYOLU (Linux):
#  ─────────────────────────────────────────────────────────────────────────
#  → Masaüstüne  seymen_raporlama.desktop  dosyası oluşturun (aşağıda şablon)
#
#  [Desktop Entry]
#  Name=Seymen Raporlama
#  Exec=python3 /tam/yol/seymen_raporlama.py
#  Icon=/tam/yol/SeymenRaporlama/assets/icon.png
#  Type=Application
#  Terminal=false
#
# ============================================================================

import customtkinter as ctk
import pandas as pd
from tkinter import messagebox, ttk, filedialog
import os
import sys
import datetime

try:
    from PIL import Image, ImageTk, ImageDraw, ImageFont
    PIL_OK = True
except ImportError:
    PIL_OK = False

# ── Asset yolları ─────────────────────────────────────────────────────────────
_PROG_DIR   = os.path.dirname(os.path.abspath(__file__))
ASSETS_DIR  = os.path.join(_PROG_DIR, "assets")

def asset(dosya):
    """assets/ klasöründen dosya yolunu döndürür."""
    return os.path.join(ASSETS_DIR, dosya)

def ctk_img(dosya, w, h):
    """CTkImage döndürür; dosya yoksa None."""
    try:
        p = asset(dosya)
        if os.path.exists(p):
            img = Image.open(p).convert("RGBA")
            return ctk.CTkImage(light_image=img, dark_image=img, size=(w, h))
    except Exception:
        pass
    return None

# ── Tema modu (Dark / Light) — Varsayılan: Açık ─────────────────────────────
TEMA_MOD = "Light"
ctk.set_appearance_mode("Light")
ctk.set_default_color_theme("blue")

# ============================================================================
# UYGULAMA SABİTLERİ
# ============================================================================
UYGULAMA_ADI     = "Seymen Raporlama"
UYGULAMA_SURUMU  = "v1.0"
HAZIRLAYEN       = "S. SEYMEN"

PREGATE_EKIBI = [
    "tolga tosun", "muhammt acı", "serhat seymen",
    "muhammet malik eşber kılıç", "emre sabe", "kaan turgut", "oğuzhan öz"
]

KATEGORILER = [
    "Sorumluluk Evrakı", "Yüksekte Çalışabilir Sağlık Raporu",
    "Tank Basınç Raporu - T9", "Muayene", "Zorunlu Trafik Sigortası",
    "K1 Taşıt Kartı", "Tehlikeli Atık Sigortası", "Tank Kodu", "KKD",
    "interchange", "Şoför Kart Almaya Gelmedi",
    "Sevkiyat - Firma Sürücü İptal", "Havuz Tahliye Vanası", "İSOPA",
    "FAR-STOP-SİS FARI-LASTİK KIRIK/PATLAK",
    "Turuncu Plaka ve İkaz Tabelası", "Yakıt Deposu Sızıntı", "Diğer"
]

# ── Pregate sunum — yeni red sınıflandırması (İptal Adımı sütunundan) ─────────
ANA_NEDENLER = [
    "EVRAKSAL", "FİRMA TALEBİ İPTAL", "SÜRÜCÜ", "DONANIMSAL", "KKD", "DİĞER"
]

PREGATE_RED_ADIMLARI   = {"PregateRed", "PregateKayitRed"}
SEVKIYAT_RED_ADIMLARI  = {"SevkiyatPlanlama"}
OPERASYON_RED_ADIMLARI = {"PlatformRed", "TerminalRed"}

TUM_RED_ADIMLARI = PREGATE_RED_ADIMLARI | SEVKIYAT_RED_ADIMLARI | OPERASYON_RED_ADIMLARI

# ============================================================================
# RENK PALETİ  —  Koyu / Açık tema
# ============================================================================
def _palet_hesapla(mod="Dark"):
    if mod == "Dark":
        return {
            "bg":           "#0A0E17",
            "sidebar":      "#0F1520",
            "card":         "#141C2B",
            "card2":        "#1A2438",
            "border":       "#252D3D",
            "hover":        "#1E2A3F",
            "altin":        "#D4AF37",
            "altin_dim":    "#8B7222",
            "mavi":         "#2E86DE",
            "yesil":        "#26A65B",
            "kirmizi":      "#C0392B",
            "turuncu":      "#E67E22",
            "mor":          "#8E44AD",
            "camgobegi":    "#16A085",
            "metin":        "#E8EDF5",
            "metin_dim":    "#7F8C9A",
            "metin_koyu":   "#4A5568",
            "kimya":        "#7B241C",
            "dow":          "#154360",
            "terminal":     "#0E6251",
            "antrepo":      "#4A235A",
            "ek":           "#6E2F1A",
            # metin tag renkleri
            "tag_tek_bg":   "#141C2B",
            "tag_cift_bg":  "#111827",
            "tag_h1_bg":    "#080C14",
        }
    else:  # Light — beyaz içerik, koyu mavi sidebar
        return {
            "bg":           "#F0F4F8",
            "sidebar":      "#1B3A5C",
            "card":         "#FFFFFF",
            "card2":        "#EBF2FA",
            "border":       "#C8D8EA",
            "hover":        "#D6E8F5",
            "altin":        "#D4AF37",
            "altin_dim":    "#A07820",
            "mavi":         "#1A6DB5",
            "yesil":        "#1A7A42",
            "kirmizi":      "#9B2335",
            "turuncu":      "#C05A0A",
            "mor":          "#6B3590",
            "camgobegi":    "#0D7A63",
            "metin":        "#1A202C",
            "metin_dim":    "#4A5568",
            "metin_koyu":   "#718096",
            "kimya":        "#922B21",
            "dow":          "#1A5276",
            "terminal":     "#0E6655",
            "antrepo":      "#6C3483",
            "ek":           "#935116",
            "tag_tek_bg":   "#F7FAFC",
            "tag_cift_bg":  "#EDF2F7",
            "tag_h1_bg":    "#E2EEF9",
        }

C = _palet_hesapla("Light")  # Varsayılan: Açık tema

# ============================================================================
# YARDIMCI FONKSİYONLAR
# ============================================================================
def sutun_bul(df, arananlar):
    """DataFrame içinde anahtar kelimelerle sütun adı arar."""
    for k in arananlar:
        for c in df.columns:
            if k.lower() in str(c).lower():
                return c
    return None


def tr_normalize(metin):
    """Türkçe karakterleri normalize eder (arama için)."""
    return (str(metin).lower()
            .replace("i̇", "i").replace("ı", "i")
            .replace("ş", "s").replace("ğ", "g")
            .replace("ü", "u").replace("ö", "o")
            .replace("ç", "c"))


def yapay_zeka_kategori(metin):
    """Açıklama metninden otomatik red kategorisi çıkarır."""
    m = tr_normalize(metin)
    if any(k in m for k in ["sorumluluk", "taahhutname"]):
        return "Sorumluluk Evrakı"
    if any(k in m for k in ["yuksek", "saglik", "rapor"]):
        return "Yüksekte Çalışabilir Sağlık Raporu"
    if any(k in m for k in ["basinc", "t9"]):
        return "Tank Basınç Raporu - T9"
    if any(k in m for k in ["muayene", "vize"]):
        return "Muayene"
    if any(k in m for k in ["trafik", "sigorta"]) and "atik" not in m:
        return "Zorunlu Trafik Sigortası"
    if "k1" in m:
        return "K1 Taşıt Kartı"
    if "tehlikeli atik" in m:
        return "Tehlikeli Atık Sigortası"
    if "tank kodu" in m:
        return "Tank Kodu"
    if any(k in m for k in ["kkd","baret","gozluk","kemer","reflektor","ayakkabi"]):
        return "KKD"
    if "interchange" in m:
        return "interchange"
    if any(k in m for k in ["kart almaya gelmedi","kart almadi"]):
        return "Şoför Kart Almaya Gelmedi"
    if any(k in m for k in ["firma iptal","surucu iptal","sevkiyat iptal","gelmedi","arac arizasi"]):
        return "Sevkiyat - Firma Sürücü İptal"
    if any(k in m for k in ["havuz","tahliye vanasi","vana"]):
        return "Havuz Tahliye Vanası"
    if "isopa" in m:
        return "İSOPA"
    if any(k in m for k in ["far","stop","sis","lastik","kirik","patlak","lamba","fari"]):
        return "FAR-STOP-SİS FARI-LASTİK KIRIK/PATLAK"
    if any(k in m for k in ["turuncu plaka","ikaz","tabela","levha","plaka"]):
        return "Turuncu Plaka ve İkaz Tabelası"
    if any(k in m for k in ["sizinti","kacak","yakit"]):
        return "Yakıt Deposu Sızıntı"
    return "Diğer"


# ============================================================================
# VERİ MOTORU  —  tüm hesaplama mantığı burada
# ============================================================================
class VeriMotoru:
    """
    Tüm analiz & hesaplama işlemleri bu sınıfta.
    Arayüz kodundan tamamen bağımsız tutulmuştur.
    """

    def __init__(self):
        self.df_dinamik = pd.DataFrame()
        self.df_pregate = pd.DataFrame()
        self.df_ek      = pd.DataFrame()
        self.df_pol_tmp = pd.DataFrame()   # Nakliyeci çalışma kopyası

        self.dosya_dinamik = ""
        self.dosya_pregate = ""
        self.dosya_ek      = ""

    # ── Yükleme ──────────────────────────────────────────────────────────────
    def _tekillesir(self, df, kaynak_adi=""):
        """
        Plaka + Kayıt Tarihi kombinasyonuna göre tekrar eden satırları çıkarır.
        Farklı dosyalardan gelen aynı araç kaydını çiftlemez.
        Döner: (temizlenmiş_df, kaldırılan_satır_sayısı)
        """
        plaka_col = sutun_bul(df, ["plaka","plate","araç no","arac no"])
        tarih_col = sutun_bul(df, ["tarih","date","kayıt","giris","giriş"])

        if not plaka_col:
            return df, 0   # Plaka sütunu yoksa tekilleştirme yapma

        onceki = len(df)
        df = df.copy()
        df["_plaka_norm"] = df[plaka_col].astype(str).str.strip().str.upper().str.replace(" ","")

        if tarih_col:
            df["_tarih_norm"] = pd.to_datetime(
                df[tarih_col], errors="coerce"
            ).dt.strftime("%Y-%m-%d")
            anahtar = ["_plaka_norm","_tarih_norm"]
        else:
            anahtar = ["_plaka_norm"]

        df = df.drop_duplicates(subset=anahtar, keep="last")
        df = df.drop(columns=[c for c in ["_plaka_norm","_tarih_norm"]
                               if c in df.columns])
        return df, onceki - len(df)

    def yukle_dinamik(self, yol):
        """
        ADIM 1: Excel oku
        ADIM 2: 'Kayit Yapan Lokasyon' sutunundan 'DILOVASI FABRIKA' satirlarini sec (sabit filtre)
        ADIM 3: Filtrelenmis veriyi sakla — tekillesme yok, her kayit sayilir
        """
        df = pd.read_excel(yol)

        # 'Kayit Yapan Lokasyon' sutununu bul
        lok_col = sutun_bul(df, ["kayit yapan lokasyon", "kayit yapan",
                                  "kayıt yapan lokasyon", "lokasyon"])
        if lok_col:
            def lok_norm(s):
                return (str(s).strip().upper()
                        .replace("İ","I").replace("Ğ","G")
                        .replace("Ş","S").replace("Ü","U")
                        .replace("Ö","O").replace("Ç","C"))
            df = df[df[lok_col].apply(lok_norm).str.contains("DILOVASI FABRIKA", na=False)]

        self.df_dinamik    = df
        self.dosya_dinamik = yol
        return 0

    def _norm_str(self, s):
        """Türkçe karakter normalize + büyük harf."""
        return (str(s).strip().upper()
                .replace("İ","I").replace("Ğ","G")
                .replace("Ş","S").replace("Ü","U")
                .replace("Ö","O").replace("Ç","C"))

    def dinamik_kpi(self, df):
        """
        Sütun adı TAM OLARAK: 'Teknik Emniyet Cagirma'
        ONAYLANDI   → Onay
        ONAYLANMADI → Red
        Sütun adı sabit, sadece yeri değişir.
        """
        if df is None or df.empty:
            return 0, 0, 0

        # TAM sütun adıyla ara — büyük/küçük fark etmeksizin
        col = None
        for c in df.columns:
            if str(c).strip().lower() == "teknik emniyet cagirma":
                col = c
                break

        if not col:
            return len(df), 0, 0

        s = df[col].astype(str).str.strip()
        onay = int((s == "ONAYLANDI").sum())
        red  = int((s == "ONAYLANMADI").sum())
        return len(df), onay, red

    def _dinamik_filtrele(self, departman, alt_birim=None):
        """
        'Havuz Adi' sutununa gore departman ayrimi:

        Kimya  (sabit ad: 'POLISAN KIMYA DEPARTMANI')
          Alt birimler Havuz Adi'nden:
            - POLISAN KIMYA SEVKIYAT
            - POLISAN YAPIKIM SEVKIYAT

        Dow  (sabit ad: 'DOW DEPARTMANI')
          Alt birimler Havuz Adi'nden:
            - DOW CM
            - DOW PA

        Diger  (sabit: 'DIGER DEPARTMANLAR/GENEL')
          -> Kimya ve Dow listesine girmeyen TUM Havuz Adi degerleri

        alt_birim verilmisse -> sadece o havuzun satirlari doner
        """
        df = self.df_dinamik.copy()
        if df.empty:
            return df

        havuz_col = sutun_bul(df, ["havuz adi", "havuz adı", "havuz"])
        if not havuz_col:
            return df

        sh = df[havuz_col].apply(self._norm_str)

        # Kimya alt birimleri — tam ad eslesmesi
        KIMYA = {"POLISAN KIMYA SEVKIYAT", "POLISAN YAPIKIM SEVKIYAT"}
        # Dow alt birimleri — tam ad eslesmesi
        DOW   = {"DOW CM", "DOW PA"}

        km = sh.isin(KIMYA)
        dw = sh.isin(DOW)
        di = ~km & ~dw  # Geri kalan her sey -> Diger

        if   departman == "Kimya":   df_dep = df[km]
        elif departman == "Dow":     df_dep = df[dw]
        elif departman == "Diğer":   df_dep = df[di]
        else:                        return df

        # Alt birim filtresi — normalize karsilastirma
        if alt_birim and not df_dep.empty:
            mask = df_dep[havuz_col].apply(self._norm_str) == self._norm_str(alt_birim)
            df_dep = df_dep[mask]

        return df_dep

    def dinamik_havuz_detay(self, departman):
        """
        Departmandaki her benzersiz Havuz Adi icin (ad, toplam, onay, red) listesi.
        Normalize ederek tekillestirir — buyuk/kucuk ayni havuz tek gorunur.
        """
        df = self._dinamik_filtrele(departman)
        havuz_col = sutun_bul(df, ["havuz adi","havuz adı","havuz"])
        if not havuz_col or df.empty:
            return []

        gorulen = {}
        for val in df[havuz_col].dropna():
            n = self._norm_str(val)
            if n not in gorulen:
                gorulen[n] = str(val)

        sonuc = []
        for n, orijinal in sorted(gorulen.items()):
            df_p = df[df[havuz_col].apply(self._norm_str) == n]
            t, o, r = self.dinamik_kpi(df_p)
            sonuc.append((orijinal, t, o, r))
        return sonuc

    def dinamik_alt_birimler_liste(self, departman):
        """Sidebar icin — departmandaki benzersiz Havuz Adi listesi."""
        return [ad for ad, t, o, r in self.dinamik_havuz_detay(departman)]
    def yukle_pregate(self, yol):
        df = pd.read_excel(yol)
        # Tekilleştirme YOK
        self.df_pregate    = df
        self.dosya_pregate = yol
        return 0

    def yukle_ek(self, yol):
        df = pd.read_excel(yol)
        self.df_ek    = df
        self.dosya_ek = yol
        return 0

    # ── Pregate hesaplama ────────────────────────────────────────────────────
    def pregate_kpi(self, df):
        """
        Onay/Red sayımı artık 'İptal Adımı' sütunundan yapılır.
          Red  = İptal Adımı dolu (TUM_RED_ADIMLARI içinde)
          Onay = İptal Adımı boş
        """
        if df is None or df.empty:
            return 0, 0, 0
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        if not adim_col:
            # Eski davranış: Durum sütununa geri dön
            durum_col = sutun_bul(df, ["durum"])
            if not durum_col:
                return len(df), 0, 0
            sd = df[durum_col].astype(str).str.lower().str.strip()
            onay = int((sd == "onay").sum())
            red  = int(sd.isin(["pregate red","sevkiyat red","operasyon red"]).sum())
            return len(df), onay, red
        sa = df[adim_col].astype(str).str.strip()
        red  = int(sa.isin(TUM_RED_ADIMLARI).sum())
        onay = int(len(df)) - red
        return len(df), onay, red

    def pregate_red_detay_liste(self, departman):
        """
        PREGATE RED / SEVKİYAT RED / OPERASYON RED adet sözlüğü — İptal Adımı'ndan.
        Departman ayrımı zorunludur; farklı departmanlar asla karıştırılmaz.
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return {}
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        if not adim_col:
            return {}
        sa = df[adim_col].astype(str).str.strip()
        return {
            "Pregate Red":   int(sa.isin(PREGATE_RED_ADIMLARI).sum()),
            "Sevkiyat Red":  int(sa.isin(SEVKIYAT_RED_ADIMLARI).sum()),
            "Operasyon Red": int(sa.isin(OPERASYON_RED_ADIMLARI).sum()),
        }

    def pregate_red_df(self, departman, red_tipi):
        """
        Belirtilen red grubunun DataFrame'ini döndürür — İptal Adımı'ndan.
        red_tipi: 'Pregate Red' | 'Sevkiyat Red' | 'Operasyon Red' | diğer → tümü
        Departman ayrımı zorunlu; her çağrıda departman filtreli veri gelir.
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return pd.DataFrame()
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        if not adim_col:
            return pd.DataFrame()
        sa = df[adim_col].astype(str).str.strip()
        if red_tipi == "Pregate Red":
            return df[sa.isin(PREGATE_RED_ADIMLARI)].copy()
        elif red_tipi == "Sevkiyat Red":
            return df[sa.isin(SEVKIYAT_RED_ADIMLARI)].copy()
        elif red_tipi == "Operasyon Red":
            return df[sa.isin(OPERASYON_RED_ADIMLARI)].copy()
        else:
            return df[sa.isin(TUM_RED_ADIMLARI)].copy()

    def pregate_red_detay(self, departman):
        """
        PREGATE / SEVKİYAT / OPERASYON RED — İptal Adımı sütunundan.
        Departman ayrımı zorunlu; veriler asla karıştırılmaz.
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return {}, [], []
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        if not adim_col:
            return {}, [], []
        sa = df[adim_col].astype(str).str.strip()

        pg_mask  = sa.isin(PREGATE_RED_ADIMLARI)
        sev_mask = sa.isin(SEVKIYAT_RED_ADIMLARI)
        op_mask  = sa.isin(OPERASYON_RED_ADIMLARI)

        kategoriler = {
            "Pregate Red":   int(pg_mask.sum()),
            "Sevkiyat Red":  int(sev_mask.sum()),
            "Operasyon Red": int(op_mask.sum()),
        }
        red_df = df[pg_mask | sev_mask | op_mask]
        nak_col = sutun_bul(red_df, ["nakliyeci adi","nakliyeci adı","nakliye","Nakliyeci"])
        top_nak = []
        if nak_col:
            s = red_df[nak_col].astype(str).str.strip()
            top_nak = list(s[s.str.lower() != "nan"].value_counts().head(10).items())
        # Ana Neden varsa önce onu kullan, yoksa eski açıklama sütununa bak
        neden_col = sutun_bul(red_df, ["ana neden","ana_neden"])
        if not neden_col:
            neden_col = sutun_bul(red_df, ["iptal aciklama","aciklama","neden"])
        top_acik = []
        if neden_col:
            s = red_df[neden_col].astype(str).str.strip()
            top_acik = list(s[s.str.lower() != "nan"].value_counts().head(10).items())
        return kategoriler, top_nak, top_acik

    def pregate_alt_birim_detay(self, departman, red_tipi):
        """
        red_tipi: 'Pregate Red' | 'Sevkiyat Red' | 'Operasyon Red'
        İptal Adımı'ndan filtreler; departman ayrımı zorunlu.
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return [], []
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        iptal_col = sutun_bul(df, ["iptal eden","İptal Eden","eden"])
        if not adim_col:
            return [], []
        sa = df[adim_col].astype(str).str.strip()
        if red_tipi == "Pregate Red":
            maske = sa.isin(PREGATE_RED_ADIMLARI)
        elif red_tipi == "Sevkiyat Red":
            maske = sa.isin(SEVKIYAT_RED_ADIMLARI)
        elif red_tipi == "Operasyon Red":
            maske = sa.isin(OPERASYON_RED_ADIMLARI)
        else:
            maske = sa.isin(TUM_RED_ADIMLARI)
        df_fil = df[maske]
        nak_col = sutun_bul(df_fil, ["nakliyeci adi","nakliyeci adı","nakliye","Nakliyeci"])
        firma = []
        if nak_col:
            s = df_fil[nak_col].astype(str).str.strip()
            firma = list(s[s.str.lower() != "nan"].value_counts().head(15).items())
        kisi = []
        if iptal_col:
            s = df_fil[iptal_col].astype(str).str.strip()
            kisi = list(s[s.str.lower() != "nan"].value_counts().head(10).items())
        return firma, kisi

    def _pregate_filtrele(self, departman):
        df = self.df_pregate.copy()
        if df.empty:
            return df
        dep_col = sutun_bul(df, ["departman"])
        if dep_col:
            sd = df[dep_col].astype(str).str.lower()
            if departman == "Poliport Terminal":
                return df[sd.str.contains("terminal", na=False)]
            elif departman == "Antrepo":
                return df[sd.str.contains("kuru|antrepo", na=False)]
        return df

    # ── Nakliyeci Matrix (Ana Neden × Nakliyeci pivot) ───────────────────────
    def nakliyeci_matrix(self, departman="Poliport Terminal"):
        """
        Nakliyeci Red Matrix — her zaman departman bazlı çalışır.

        Sütun yapısı:
          5 × Ana Neden  |  PREGATE RED  |  SEVKİYAT RED  |  OPERASYON RED
          |  Toplam Red  |  Toplam Giriş

        Satır: Nakliyeci firma adı.  Son satır: TOPLAM.

        departman: 'Poliport Terminal' (varsayılan) veya 'Antrepo'
        Farklı departmanlar asla karıştırılmaz.
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return pd.DataFrame()

        nak_col  = sutun_bul(df, ["nakliyeci adi","nakliyeci adı","nakliye","Nakliyeci"])
        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        neden_col = sutun_bul(df, ["ana neden","Ana Neden","ana_neden"])

        if not nak_col or not adim_col:
            return pd.DataFrame()

        df = df.copy()
        df["_N"] = df[nak_col].astype(str).str.strip().str.upper()
        df = df[df["_N"] != "NAN"]

        # Tüm nakliyeci girişleri (onay + red)
        toplam_giris = df["_N"].value_counts()

        # Sadece red satırları
        sa = df[adim_col].astype(str).str.strip()
        red_df = df[sa.isin(TUM_RED_ADIMLARI)].copy()

        if red_df.empty:
            return pd.DataFrame()

        sa_red = red_df[adim_col].astype(str).str.strip()

        # ── Ana Neden pivot ──────────────────────────────────────────────────
        if neden_col:
            red_df["_ANA_NEDEN"] = red_df[neden_col].astype(str).str.strip().str.upper()
            # Listede olmayan değerleri DİĞER'e aktar — toplam tutarlılığı için
            _bilinen = set(ANA_NEDENLER)
            red_df["_ANA_NEDEN"] = red_df["_ANA_NEDEN"].apply(
                lambda x: x if x in _bilinen else "DİĞER"
            )
        else:
            red_df["_ANA_NEDEN"] = "DİĞER"

        pivot = pd.crosstab(red_df["_N"], red_df["_ANA_NEDEN"])

        # Tüm Ana Neden sütunlarını ekle (yoksa 0)
        for k in ANA_NEDENLER:
            if k not in pivot.columns:
                pivot[k] = 0

        # Sadece ANA_NEDENLER sütunlarını al (sıralı)
        mevcut = [k for k in ANA_NEDENLER if k in pivot.columns]
        pivot = pivot.reindex(columns=mevcut, fill_value=0)

        # ── Red grubu toplam sütunları ───────────────────────────────────────
        def _red_grup_say(maske_set):
            """Her nakliyeci için belirli red grubundaki kayıt sayısı."""
            alt = red_df[sa_red.isin(maske_set)].copy()
            if alt.empty:
                return pd.Series(0, index=pivot.index)
            return alt.groupby("_N").size().reindex(pivot.index, fill_value=0)

        pivot["PREGATE RED"]   = _red_grup_say(PREGATE_RED_ADIMLARI)
        pivot["SEVKİYAT RED"]  = _red_grup_say(SEVKIYAT_RED_ADIMLARI)
        pivot["OPERASYON RED"] = _red_grup_say(OPERASYON_RED_ADIMLARI)

        # ── Özet sütunlar ────────────────────────────────────────────────────
        pivot["Toplam Red"]   = pivot[["PREGATE RED","SEVKİYAT RED","OPERASYON RED"]].sum(axis=1)
        pivot["Toplam Giriş"] = pivot.index.map(toplam_giris).fillna(0).astype(int)

        pivot = pivot.sort_values("Toplam Red", ascending=False)
        pivot.loc["TOPLAM"] = pivot.sum()
        pivot.index.name = "Firma Adı"
        return pivot

    # ── Red tipi bazlı nakliyeci listeleri (Excel export için) ───────────────
    def red_tipi_nakliyeci_listesi(self, departman="Poliport Terminal"):
        """
        PREGATE RED / SEVKİYAT RED / OPERASYON RED listelerini ayrı DataFrame olarak döndürür.
        İptal Adımı sütunundan filtreler. Departman ayrımı zorunludur.

        Dönen dict anahtarları:
          pregate_red    : PregateRed + PregateKayitRed kayıtları
          sevkiyat_red   : SevkiyatPlanlama kayıtları
          operasyon_red  : PlatformRed + TerminalRed kayıtları
          genel_matrix   : Nakliyeci × Ana Neden pivot (+ red grubu toplamları)
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return {}

        adim_col = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        nak_col  = sutun_bul(df, ["nakliyeci adi","nakliyeci adı","nakliye","Nakliyeci"])

        if not adim_col:
            return {}

        df = df.copy()
        sa = df[adim_col].astype(str).str.strip()

        if nak_col:
            df["Nakliyeci_Temiz"] = df[nak_col].astype(str).str.strip().str.upper()

        def _hazirla(maske, etiket):
            alt = df[maske].copy()
            alt.insert(0, "Red Tipi", etiket)
            return alt

        pregate_df   = _hazirla(sa.isin(PREGATE_RED_ADIMLARI),   "Pregate Red")
        sevkiyat_df  = _hazirla(sa.isin(SEVKIYAT_RED_ADIMLARI),  "Sevkiyat Red")
        operasyon_df = _hazirla(sa.isin(OPERASYON_RED_ADIMLARI), "Operasyon Red")

        genel_pivot = self.nakliyeci_matrix(departman)

        return {
            "pregate_red":   pregate_df,
            "sevkiyat_red":  sevkiyat_df,
            "operasyon_red": operasyon_df,
            "genel_matrix":  genel_pivot,
        }


    # ── İptal Açıklama bazlı detay red matrix ────────────────────────────────
    def detay_red_matrix(self, departman="Poliport Terminal"):
        """
        Nakliyeci × İptal Açıklama pivot — her red tipi için ayrı DataFrame.

        Ana Neden yerine gerçek iptal açıklaması (serbest metin) kullanılır.
        Her DataFrame: satır = Nakliyeci, sütun = İptal Açıklama değerleri,
        son sütun = Toplam. Son satır = TOPLAM.

        Döner: dict {"pregate": df, "sevkiyat": df, "operasyon": df}
        """
        df = self._pregate_filtrele(departman)
        if df.empty:
            return {}

        nak_col  = sutun_bul(df, ["nakliyeci adi","nakliyeci adı","nakliye","Nakliyeci"])
        adim_col = sutun_bul(df, ["iptal adimi","iptal adımı","İptal Adımı"])
        acik_col = sutun_bul(df, ["iptal aciklama","İptal Aciklama",
                                   "iptal açıklama","İptal Açıklama",
                                   "aciklama","açıklama","neden"])

        if not nak_col or not adim_col:
            return {}

        df = df.copy()
        df["_N"] = df[nak_col].astype(str).str.strip().str.upper()
        df = df[df["_N"] != "NAN"]

        sa = df[adim_col].astype(str).str.strip()

        if acik_col:
            df["_DETAY"] = df[acik_col].astype(str).str.strip()
        else:
            df["_DETAY"] = "(Açıklama yok)"

        # Boş/nan detayları etiketle
        df["_DETAY"] = df["_DETAY"].replace({"nan": "(Açıklama yok)", "": "(Açıklama yok)"})

        def _pivot_yap(maske):
            alt = df[maske].copy()
            if alt.empty:
                return pd.DataFrame()
            pivot = pd.crosstab(alt["_N"], alt["_DETAY"])
            pivot["Toplam"] = pivot.sum(axis=1)
            pivot = pivot.sort_values("Toplam", ascending=False)
            pivot.loc["TOPLAM"] = pivot.sum()
            pivot.index.name = "Nakliyeci"
            return pivot

        return {
            "pregate":   _pivot_yap(sa.isin(PREGATE_RED_ADIMLARI)),
            "sevkiyat":  _pivot_yap(sa.isin(SEVKIYAT_RED_ADIMLARI)),
            "operasyon": _pivot_yap(sa.isin(OPERASYON_RED_ADIMLARI)),
        }

    # ── Tüm departmanlar için özet sayım tablosu ─────────────────────────────
    def ozet_sayim_tablosu(self):
        """Ana sayfa KPI'larını DataFrame satırları olarak döndürür."""
        satirlar = []
        for dep, dep_adi, tip in [
            ("Kimya",  "Polisan Kimya Departmanı", "Dinamik / Fabrika"),
            ("Dow",    "Dow Departmanı",            "Dinamik / Fabrika"),
            ("Diğer",  "Diğer / Genel",             "Dinamik / Fabrika"),
            ("Poliport Terminal", "Poliport Terminal", "Pregate / Terminal"),
            ("Antrepo",           "Antrepo",           "Pregate / Terminal"),
        ]:
            if tip.startswith("Dinamik"):
                t, o, r = self.dinamik_kpi(self._dinamik_filtrele(dep))
            else:
                t, o, r = self.pregate_kpi(self._pregate_filtrele(dep))
            oran = f"%{round(o/t*100,1)}" if t > 0 else "—"
            satirlar.append({
                "Departman":   dep_adi,
                "Tip":         tip,
                "Toplam Araç": t,
                "Onaylanan":   o,
                "Reddedilen":  r,
                "Onay Oranı":  oran,
            })
        return pd.DataFrame(satirlar)

    # ── Eksik veri analizi ───────────────────────────────────────────────────
    def eksik_analiz(self, df, sutunlar):
        sonuc = {}
        for s in sutunlar:
            col = sutun_bul(df, [s])
            if col:
                bos = int(df[col].isna().sum()
                          + (df[col].astype(str).str.strip() == "").sum())
                if bos > 0:
                    sonuc[col] = bos
        return sonuc

    # ── Araç / Plaka Arama Motoru ────────────────────────────────────────────
    def arac_ara(self, sorgu, filtreler=None):
        """
        Sorguya göre tüm yüklü dosyalarda arama yapar.
        filtreler: dict  {
            "durum": "Onay"|"Red"|"Tümü",
            "tarih_bas": "2024-01-01" veya "",
            "tarih_bit": "2024-12-31" veya "",
        }
        Döner: {
            "dinamik":  DataFrame,   # Dinamik rapordaki eşleşmeler
            "pregate":  DataFrame,   # Pregate sunumdaki eşleşmeler
            "ek":       DataFrame,   # Ek dosyadaki eşleşmeler
        }
        Her df ayrıca "Kaynak" ve "Eşleşme Alanı" sütunu taşır.
        """
        sorgu = sorgu.strip().upper()
        if not sorgu:
            return {"dinamik": pd.DataFrame(),
                    "pregate": pd.DataFrame(),
                    "ek":      pd.DataFrame()}

        if filtreler is None:
            filtreler = {}

        # Aranacak sütun kategorileri
        PLAKA_KEYS   = ["plaka","plate","araç no","arac no","araç plaka"]
        DORSE_KEYS   = ["dorse","römork","trailer"]
        TANK_KEYS    = ["tank","konteyner","container","silo"]
        NAK_KEYS     = ["nakliyeci","nakliye","firma"]
        TARIH_KEYS   = ["tarih","date","kayıt","giriş","giris"]
        DURUM_KEYS   = ["durum","status","sonuç","sonuc","emniyet","çağrı","cagirma"]

        def _eslesir(df, kaynak_adi):
            if df is None or df.empty:
                return pd.DataFrame()

            df = df.copy()

            # Arama yapılacak sütunları bul
            arama_sutunlari = {}
            for kategori, anahtar_listesi in [
                ("Plaka",    PLAKA_KEYS),
                ("Dorse",    DORSE_KEYS),
                ("Tank No",  TANK_KEYS),
                ("Nakliyeci",NAK_KEYS),
            ]:
                col = sutun_bul(df, anahtar_listesi)
                if col:
                    arama_sutunlari[kategori] = col

            if not arama_sutunlari:
                # Hiç tanınan sütun yoksa tüm metin sütunlarına bak
                for c in df.select_dtypes(include=["object"]).columns[:6]:
                    arama_sutunlari[c] = c

            # Her sütunda ara
            maskeler = []
            eslesme_labels = []
            for etiket, col in arama_sutunlari.items():
                m = df[col].astype(str).str.upper().str.contains(
                    sorgu, regex=False, na=False)
                maskeler.append(m)
                eslesme_labels.append((m, etiket))

            if not maskeler:
                return pd.DataFrame()

            toplam_maske = maskeler[0]
            for m in maskeler[1:]:
                toplam_maske = toplam_maske | m

            sonuc = df[toplam_maske].copy()
            if sonuc.empty:
                return pd.DataFrame()

            # Eşleşme alanı etiketi ekle
            eslesme_col = []
            for _, row in sonuc.iterrows():
                etiketler = []
                for m, et in eslesme_labels:
                    if m.loc[row.name] if row.name in m.index else False:
                        etiketler.append(et)
                eslesme_col.append(" + ".join(etiketler) if etiketler else "?")
            sonuc.insert(0, "Kaynak", kaynak_adi)
            sonuc.insert(1, "Eşleşme Alanı", eslesme_col)

            # Tarih filtresi
            tarih_col = sutun_bul(sonuc, TARIH_KEYS)
            if tarih_col and (filtreler.get("tarih_bas") or filtreler.get("tarih_bit")):
                tarih_s = pd.to_datetime(sonuc[tarih_col], errors="coerce")
                bas = filtreler.get("tarih_bas","")
                bit = filtreler.get("tarih_bit","")
                if bas:
                    try:
                        sonuc = sonuc[tarih_s >= pd.Timestamp(bas)]
                    except Exception:
                        pass
                if bit:
                    try:
                        sonuc = sonuc[tarih_s <= pd.Timestamp(bit)]
                    except Exception:
                        pass

            # Durum filtresi
            durum_fil = filtreler.get("durum","Tümü")
            if durum_fil != "Tümü":
                durum_col = sutun_bul(sonuc, DURUM_KEYS)
                if durum_col:
                    sd = sonuc[durum_col].astype(str).str.lower()
                    if durum_fil == "Onay":
                        sonuc = sonuc[
                            sd.str.contains("onay") & ~sd.str.contains("onaylanmadı|onaylanmadi|red")
                        ]
                    elif durum_fil == "Red":
                        sonuc = sonuc[
                            sd.str.contains("red|onaylanmadı|onaylanmadi")
                        ]

            return sonuc.reset_index(drop=True)

        return {
            "dinamik": _eslesir(self.df_dinamik, "Dinamik Rapor"),
            "pregate": _eslesir(self.df_pregate,  "Pregate Sunum"),
            "ek":      _eslesir(self.df_ek,        "Ek Dosya"),
        }

    # ── Genel KPI özeti ──────────────────────────────────────────────────────
    def genel_ozet(self):
        k_t,k_o,k_r   = self.dinamik_kpi(self._dinamik_filtrele("Kimya"))
        d_t,d_o,d_r   = self.dinamik_kpi(self._dinamik_filtrele("Dow"))
        di_t,di_o,di_r = self.dinamik_kpi(self._dinamik_filtrele("Diğer"))
        t_t,t_o,t_r   = self.pregate_kpi(self._pregate_filtrele("Poliport Terminal"))
        a_t,a_o,a_r   = self.pregate_kpi(self._pregate_filtrele("Antrepo"))
        gt = k_t+d_t+di_t+t_t+a_t
        go = k_o+d_o+di_o+t_o+a_o
        gr = k_r+d_r+di_r+t_r+a_r
        return {
            "toplam": gt, "onay": go, "red": gr,
            "kimya":   (k_t,k_o,k_r),
            "dow":     (d_t,d_o,d_r),
            "diger":   (di_t,di_o,di_r),
            "terminal": (t_t,t_o,t_r),
            "antrepo":  (a_t,a_o,a_r),
        }


# ============================================================================
# ANA PENCERE
# ============================================================================
class SeymenRaporlama(ctk.CTk):

    def __init__(self):
        super().__init__()

        # Pencere ayarları
        self.title(f"{UYGULAMA_ADI}  —  {HAZIRLAYEN}  |  {UYGULAMA_SURUMU}")
        self.geometry("1500x920")
        self.minsize(1200, 760)
        self.configure(fg_color=C["bg"])

        # ── Pencere ikonu (Windows .ico / macOS PNG) ─────────────────────────
        try:
            import platform
            if platform.system() == "Windows":
                ico = asset("SYMN_Pregate_Raporlama.ico")
                if not os.path.exists(ico):
                    ico = asset("logo.ico")
                if os.path.exists(ico):
                    self.iconbitmap(ico)
            else:
                # macOS / Linux — PhotoImage ile PNG kullan
                for png_name in ["logo_256.png", "logo_sidebar.png", "logo.png"]:
                    p = asset(png_name)
                    if os.path.exists(p) and PIL_OK:
                        from PIL import ImageTk
                        _icon_img = ImageTk.PhotoImage(
                            Image.open(p).resize((64, 64), Image.LANCZOS))
                        self.wm_iconphoto(True, _icon_img)
                        self._icon_ref = _icon_img  # GC koruması
                        break
        except Exception:
            pass  # İkon yüklenemezse program çalışmaya devam eder

        # Görsel varlıkları yükle
        self._img_logo        = ctk_img("logo_sidebar.png",   160, 60)
        self._img_logo_kucuk  = ctk_img("logo.png",            36, 36)
        self._img_poliport    = ctk_img("poliport.jpg",        560, 220)
        self._img_kuru        = ctk_img("kuru_yuk.jpg",        560, 220)
        self._img_footer      = ctk_img("footer_small.png",   220, 70)
        self._img_login_bg    = ctk_img("login_bg_resized.jpg", 1500, 920)

        # Veri motoru
        self.motor = VeriMotoru()

        # Izgara
        self.grid_rowconfigure(0, weight=1)
        self.grid_columnconfigure(1, weight=1)

        # Splash ekranını göster, sonra ana arayüzü yükle
        self._splash_goster()

    # =========================================================================
    # SPLASH / GİRİŞ EKRANI
    # =========================================================================
    def _splash_goster(self):
        """Program açılırken gösterilen tanıtım ekranı."""
        # Tüm ekranı kapla
        self.splash = ctk.CTkFrame(self, fg_color="#0A0E17", corner_radius=0)
        self.splash.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Arka plan görsel — login_bg
        if self._img_login_bg:
            bg_lbl = ctk.CTkLabel(self.splash, image=self._img_login_bg, text="")
            bg_lbl.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Koyu overlay
        overlay = ctk.CTkFrame(self.splash, fg_color="#111111", corner_radius=0)
        overlay.place(relx=0, rely=0, relwidth=1, relheight=1)

        # Merkez kart
        kart = ctk.CTkFrame(
            self.splash,
            fg_color="#0D1420",
            corner_radius=18,
            border_width=1,
            border_color="#8B7222",
            width=560,
            height=440
        )
        kart.place(relx=0.5, rely=0.5, anchor="center")
        kart.grid_propagate(False)
        kart.pack_propagate(False)

        # Logo
        if self._img_logo:
            ctk.CTkLabel(kart, image=self._img_logo, text="").pack(pady=(34, 0))
        else:
            ctk.CTkLabel(kart, text="SR",
                         font=("Arial", 52, "bold"),
                         text_color="#D4AF37").pack(pady=(34, 0))

        # Uygulama adı
        ctk.CTkLabel(
            kart,
            text=UYGULAMA_ADI,
            font=("Arial", 28, "bold"),
            text_color="#E8EDF5"
        ).pack(pady=(10, 2))

        ctk.CTkLabel(
            kart,
            text="Pregate Operasyon Yönetim Sistemi",
            font=("Arial", 13),
            text_color="#7F8C9A"
        ).pack()

        # Ayraç çizgisi
        ctk.CTkFrame(kart, height=1, fg_color="#6B5500").pack(
            fill="x", padx=40, pady=20)

        # İmza
        ctk.CTkLabel(
            kart,
            text="Tasarlayan  &  Geliştiren",
            font=("Arial", 11),
            text_color="#4A5568"
        ).pack()

        ctk.CTkLabel(
            kart,
            text=HAZIRLAYEN,
            font=("Arial", 26, "bold", "italic"),
            text_color="#D4AF37"
        ).pack(pady=(4, 0))

        ctk.CTkLabel(
            kart,
            text="Pregate Operasyon Sorumlusu  |  Poliport Terminal",
            font=("Arial", 11),
            text_color="#4A5568"
        ).pack(pady=(2, 0))

        # Yükleniyor barı
        self._splash_pb = ctk.CTkProgressBar(
            kart, height=4,
            progress_color="#D4AF37",
            fg_color="#1A2438",
            corner_radius=2
        )
        self._splash_pb.pack(fill="x", padx=40, pady=(20, 4))
        self._splash_pb.set(0)

        self._splash_lbl = ctk.CTkLabel(
            kart, text="Yükleniyor...",
            font=("Arial", 11), text_color="#4A5568"
        )
        self._splash_lbl.pack()

        # Versiyon
        ctk.CTkLabel(
            kart,
            text=f"{UYGULAMA_SURUMU}   ©  {datetime.datetime.now().year}",
            font=("Arial", 10),
            text_color="#252D3D"
        ).pack(pady=(8, 16))

        # Animasyonlu ilerleme
        self._splash_adim = 0
        self._splash_animasyon()

    def _splash_animasyon(self):
        adimlar = [
            (0.15, "Arayüz hazırlanıyor..."),
            (0.40, "Bileşenler yükleniyor..."),
            (0.65, "Veri motoru başlatılıyor..."),
            (0.85, "Son dokunuşlar..."),
            (1.00, "Hazır!"),
        ]
        if self._splash_adim < len(adimlar):
            oran, mesaj = adimlar[self._splash_adim]
            self._splash_pb.set(oran)
            self._splash_lbl.configure(text=mesaj)
            self._splash_adim += 1
            self.after(320, self._splash_animasyon)
        else:
            self.after(400, self._splash_kapat)

    def _splash_kapat(self):
        self.splash.destroy()
        # Ana arayüzü kur
        self._ttk_stili_kur()
        self._sidebar_kur()
        self._ana_alan_kur()
        self._hosgeldin_goster()
        # Tüm Treeview widget'larına global Ctrl+C / Cmd+C bağla
        self._global_copy_bind()

    # =========================================================================
    # TTK STİL (Treeview için)
    # =========================================================================
    def _ttk_stili_kur(self):
        s = ttk.Style()
        s.theme_use("default")
        s.configure("SR.Treeview",
                    background="#141C2B", foreground=C["metin"],
                    rowheight=30, fieldbackground="#141C2B",
                    borderwidth=0, font=("Consolas", 11))
        s.map("SR.Treeview",
              background=[("selected", "#1E3A5F")])
        s.configure("SR.Treeview.Heading",
                    background="#0A0E17", foreground=C["altin"],
                    font=("Arial", 11, "bold"), relief="flat", padding=(8,6))
        s.map("SR.Treeview.Heading",
              background=[("active", "#1A2438")])
        s.configure("SR.Scrollbar",
                    background=C["border"], troughcolor=C["card"],
                    relief="flat", borderwidth=0)

    # =========================================================================
    # GLOBAL KOPYALAMA / SEÇİM  —  tüm Treeview ve metin alanları
    # =========================================================================
    def _global_copy_bind(self):
        """
        Uygulama genelinde kopyala/seç:
        - Tüm Treeview'lere Ctrl+C / Cmd+C → seçili satırları TSV kopyalar.
        - Tüm Entry/CTkEntry alanlarına Ctrl+A → tümünü seç.
        - Mevcut txt_rapor textbox zaten kopyalanabilir (CTkTextbox).
        """
        def _treeview_kopyala(event):
            w = event.widget
            if not isinstance(w, ttk.Treeview):
                return
            try:
                secili = w.selection()
                cols = w["columns"]
                basliklar = [w.heading(c)["text"] for c in cols]
                satirlar  = ["\t".join(basliklar)]
                items = secili if secili else w.get_children()
                for item in items:
                    vals = [str(v) for v in w.item(item, "values")]
                    satirlar.append("\t".join(vals))
                self.clipboard_clear()
                self.clipboard_append("\n".join(satirlar))
                self._bildirim(f"✔  {len(satirlar)-1} satır kopyalandı.")
            except Exception:
                pass

        # Tüm Treeview widget sınıfına bağla (tüm örnekler yakalanır)
        self.bind_class("Treeview", "<Control-c>", _treeview_kopyala, add=True)
        self.bind_class("Treeview", "<Command-c>",  _treeview_kopyala, add=True)

        # Tüm Text (ve CTkTextbox içindeki Text) alanları seçilebilir — tkinter varsayılan
        # CTkTextbox içindeki _textbox Text widget'ı zaten seçilebilir.
        # Entry için Ctrl+A → tümünü seç
        def _entry_tumu_sec(event):
            w = event.widget
            try:
                w.select_range(0, "end")
                w.icursor("end")
            except Exception:
                pass
        self.bind_class("Entry",    "<Control-a>", _entry_tumu_sec, add=True)
        self.bind_class("Entry",    "<Command-a>",  _entry_tumu_sec, add=True)

    # =========================================================================
    # SIDEBAR
    # =========================================================================
    def _sidebar_kur(self):
        self.sidebar = ctk.CTkScrollableFrame(
            self, width=258, corner_radius=0,
            fg_color=C["sidebar"],
            scrollbar_button_color="#2D5A8E",
            scrollbar_button_hover_color="#3A72AE"
        )
        self.sidebar.grid(row=0, column=0, sticky="nsew")

        # ── Logo / Başlık — her temada beyaz arka plan ───────────────────────
        logo_satir = ctk.CTkFrame(self.sidebar, fg_color="#FFFFFF",
                                   corner_radius=0, height=60)
        logo_satir.pack(fill="x")
        logo_satir.pack_propagate(False)

        # logo.png — beyaz arka planlı Poliport logosu
        img_k = ctk_img("logo.png", 120, 44) or ctk_img("poliport.jpg", 120, 44)
        if img_k:
            ctk.CTkLabel(logo_satir, image=img_k, text="",
                         fg_color="#FFFFFF"
                         ).pack(side="left", padx=(10,6), pady=8)
        else:
            ctk.CTkLabel(logo_satir, text="Poliport",
                         font=("Arial", 16, "bold"),
                         text_color="#003366",
                         fg_color="#FFFFFF"
                         ).pack(side="left", padx=(10,6), pady=8)

        ctk.CTkLabel(logo_satir,
                     text=f"Pregate  {UYGULAMA_SURUMU}",
                     font=("Arial", 9),
                     text_color="#334E68",
                     fg_color="#FFFFFF",
                     anchor="w"
                     ).pack(side="left", pady=8)

        self._ayrac()

        # ── DOSYA YÖNETİMİ — tek buton ───────────────────────────────────────
        self._grup_basligi("📁  DOSYA YÖNETİMİ")

        ctk.CTkButton(
            self.sidebar,
            text="📂  Dosya Yükle / Yönet",
            fg_color="#2D5A8E",
            hover_color="#1B3A5C",
            font=("Arial", 11, "bold"),
            height=30, corner_radius=6,
            command=self._dosya_popup
        ).pack(pady=(2, 1), padx=14, fill="x")

        self._lbl_dosya_ozet = ctk.CTkLabel(
            self.sidebar,
            text="  Henüz dosya yüklenmedi",
            font=("Arial", 9),
            text_color="#7AAED0",
            anchor="w"
        )
        self._lbl_dosya_ozet.pack(anchor="w", padx=14, pady=(0, 2))

        self._ayrac()

        # ── ANA SAYFA ─────────────────────────────────────────────────────────
        ctk.CTkButton(
            self.sidebar,
            text="🏠  ANA SAYFA",
            fg_color=C["turuncu"],
            hover_color="#B7600D",
            font=("Arial", 12, "bold"),
            height=34, corner_radius=6,
            command=self._ana_sayfa
        ).pack(pady=(4, 2), padx=14, fill="x")

        self._ayrac()

        # ── ARAÇ / PLAKA ARAMA ────────────────────────────────────────────────
        ctk.CTkButton(
            self.sidebar,
            text="🔍  Araç / Plaka Ara",
            fg_color="#2D5A8E",
            hover_color="#1B3A5C",
            font=("Arial", 11, "bold"),
            height=30, corner_radius=6,
            command=self._goster_arama
        ).pack(pady=(2, 2), padx=14, fill="x")

        self._ayrac()

        # ── PREGATE SUNUM — ÖNCE ─────────────────────────────────────────────
        self._grup_basligi("⚓  PREGATE SUNUM  —  TERMİNAL")

        self._dep_btn_ekle(
            "🏗  Poliport Terminal", C["terminal"],
            lambda: self._goster_pregate("Poliport Terminal"),
            [("Pregate Red",   None, "Pregate Red"),
             ("Sevkiyat Red",  None, "Sevkiyat Red"),
             ("Operasyon Red", None, "Operasyon Red")],
            "#0C3028", pregate_alt=True
        )
        self._menu_btn("📦  Antrepo / Kuru Yük",
                       C["antrepo"],
                       lambda: self._goster_pregate("Antrepo"))

        ctk.CTkButton(
            self.sidebar,
            text="🚛  Nakliyeci Analizi",
            fg_color="#1B5E8B", hover_color="#14476B",
            font=("Arial", 11, "bold"),
            height=30, corner_radius=6,
            command=self._goster_nakliyeci
        ).pack(pady=(2, 2), padx=14, fill="x")

        self._ayrac()

        # ── DİNAMİK RAPOR — SONRA ────────────────────────────────────────────
        self._grup_basligi("🏭  DİNAMİK RAPOR  —  FABRİKA")

        # Kimya — sadece Excel'deki 2 havuz (normalize tekilleştirilmiş)
        kimya_altlar = []
        if not self.motor.df_dinamik.empty:
            for havuz in self.motor.dinamik_alt_birimler_liste("Kimya"):
                kimya_altlar.append((havuz, "Kimya", havuz))

        self._dep_btn_ekle(
            "⚗  Polisan Kimya Departmanı", C["kimya"],
            lambda: self._goster_dinamik("Kimya"),
            kimya_altlar, "#3D100C"
        )

        # Dow — sadece Excel'deki havuzlar
        dow_altlar = []
        if not self.motor.df_dinamik.empty:
            for havuz in self.motor.dinamik_alt_birimler_liste("Dow"):
                dow_altlar.append((havuz, "Dow", havuz))
        self._dep_btn_ekle(
            "🧪  Dow Departmanı", C["dow"],
            lambda: self._goster_dinamik("Dow"),
            dow_altlar, "#082030"
        )

        # Diğer — sadece Excel'deki havuzlar
        diger_altlar = []
        if not self.motor.df_dinamik.empty:
            for havuz in self.motor.dinamik_alt_birimler_liste("Diğer"):
                diger_altlar.append((havuz, "Diğer", havuz))
        self._dep_btn_ekle(
            "🔩  Diğer / Genel", "#5D4037",
            lambda: self._goster_dinamik("Diğer"),
            diger_altlar, "#2C1A0E"
        )

        self._ayrac()

        # ── ÇIKTI ─────────────────────────────────────────────────────────────
        self._grup_basligi("💾  ÇIKTI & RAPORLAMA")

        ctk.CTkButton(
            self.sidebar,
            text="📊  Excel Raporu Oluştur",
            fg_color="#7D6608", hover_color="#5D4E08",
            text_color=C["altin"], font=("Arial", 11, "bold"),
            height=30, corner_radius=6,
            command=self._export_excel
        ).pack(pady=2, padx=14, fill="x")

        ctk.CTkLabel(
            self.sidebar,
            text="  Pregate / Sevkiyat / Operasyon Red + Tüm",
            font=("Arial", 8), text_color="#7AAED0",
            anchor="w"
        ).pack(anchor="w", padx=14, pady=(0,2))

        # ── İmza ─────────────────────────────────────────────────────────────
        ctk.CTkFrame(self.sidebar, height=1,
                     fg_color="#2D5A8E").pack(fill="x", padx=14, pady=(8,4))
        ctk.CTkLabel(self.sidebar, text=HAZIRLAYEN,
                     font=("Arial", 11, "bold", "italic"),
                     text_color=C["altin"]).pack()
        ctk.CTkLabel(self.sidebar, text="Pregate Operasyon Sorumlusu",
                     font=("Arial", 8), text_color="#7AAED0"
                     ).pack(pady=(0, 14))
    def _dep_btn_ekle(self, etiket, ana_renk, ana_cmd,
                      alt_birimler, koyu_renk, pregate_alt=False):
        ana_btn = ctk.CTkButton(
            self.sidebar, text=etiket,
            fg_color=ana_renk, hover_color="#333",
            font=("Arial", 11), height=30,
            anchor="w", corner_radius=6,
            command=ana_cmd
        )
        ana_btn.pack(pady=(2,0), padx=14, fill="x")

        if not alt_birimler:
            return

        # "Alt Birimler" toggle oku — ana butonun HEMEN altında
        oku_btn = ctk.CTkButton(
            self.sidebar,
            text="   ▸  Alt Birimler",
            fg_color="transparent",
            hover_color=koyu_renk,
            text_color="#7AAED0",
            font=("Arial", 9), height=18, anchor="w",
            corner_radius=4,
        )
        oku_btn.pack(pady=(0,1), padx=(24,14), fill="x")

        # Alt birim çerçevesi — oku_btn'in hemen altına yerleşecek
        alt_cerceve = ctk.CTkFrame(self.sidebar, fg_color="transparent")

        for (ab_txt, dep, anahtar) in alt_birimler:
            if pregate_alt:
                cmd = lambda a=anahtar: self._goster_alt_birim(a)
            else:
                cmd = lambda d=dep, a=anahtar: self._goster_dinamik(d, a)
            kisa = ab_txt[:26] + "…" if len(ab_txt) > 26 else ab_txt
            ctk.CTkButton(
                alt_cerceve,
                text=f"   └  {kisa}",
                fg_color=koyu_renk, hover_color="#444",
                text_color="#AACCE8",
                font=("Arial", 10), height=24, anchor="w",
                corner_radius=4, command=cmd
            ).pack(pady=1, padx=(20,14), fill="x")

        durum = [False]
        oku_btn.configure(
            command=lambda d=durum, c=alt_cerceve, o=oku_btn:
                self._toggle_dep(d, c, o))

    def _toggle_dep(self, durum, cerceve, oku_btn):
        if durum[0]:
            cerceve.pack_forget()
            oku_btn.configure(text="   ▸  Alt Birimler")
            durum[0] = False
        else:
            # oku_btn'in hemen altına yerleştir
            cerceve.pack(after=oku_btn, fill="x")
            oku_btn.configure(text="   ▾  Alt Birimler")
            durum[0] = True

    def _tema_degistir(self):
        global C, TEMA_MOD
        if TEMA_MOD == "Dark":
            TEMA_MOD = "Light"
            ctk.set_appearance_mode("Light")
            C = _palet_hesapla("Light")
        else:
            TEMA_MOD = "Dark"
            ctk.set_appearance_mode("Dark")
            C = _palet_hesapla("Dark")

        # Tüm widget'ları yıkıp yeniden kur
        for w in self.winfo_children():
            try: w.destroy()
            except Exception: pass

        self.configure(fg_color=C["bg"])
        self._ttk_stili_kur()
        self._sidebar_kur()
        self._ana_alan_kur()
        self.update()
        self.update_idletasks()
        if not self.motor.df_dinamik.empty or not self.motor.df_pregate.empty:
            self._ana_sayfa()
        else:
            self._hosgeldin_goster()

    def _alt_menu_ekle(self, dep, alt_birimler, ana_renk, koyu_renk):
        """Eski metod — geriye dönük uyumluluk için boş."""
        pass

    def _toggle_fabrika_alt(self, dep, durum_attr, cerceve_attr, btn_attr):
        """Eski metod — geriye dönük uyumluluk için boş."""
        pass

    def _toggle_alt_birim(self):
        """Eski metod — geriye dönük uyumluluk için boş."""
        pass
        durum_attr  = f"_alt_ac_{dep}"
        cerceve_attr = f"_frame_alt_{dep}"
        btn_attr    = f"_btn_alt_ac_{dep}"

        setattr(self, durum_attr, False)

        btn = ctk.CTkButton(
            self.sidebar,
            text="     └  Alt Birimler  ▸",
            fg_color=koyu_renk, hover_color="#333",
            font=("Arial", 11), height=28, anchor="w",
            corner_radius=6,
            command=lambda d=dep, da=durum_attr, ca=cerceve_attr, ba=btn_attr:
                self._toggle_fabrika_alt(d, da, ca, ba)
        )
        btn.pack(pady=1, padx=18, fill="x")
        setattr(self, btn_attr, btn)

        cerceve = ctk.CTkFrame(self.sidebar, fg_color="transparent")
        setattr(self, cerceve_attr, cerceve)

        for etiket, anahtar in alt_birimler:
            ctk.CTkButton(
                cerceve,
                text=f"        •  {etiket}",
                fg_color=koyu_renk, hover_color="#444",
                font=("Arial", 11), height=26, anchor="w",
                corner_radius=5,
                command=lambda d=dep, a=anahtar: self._goster_dinamik(d, a)
            ).pack(pady=1, padx=4, fill="x")

    def _toggle_fabrika_alt(self, dep, durum_attr, cerceve_attr, btn_attr):
        """Fabrika departmanı alt birim menüsünü aç/kapa."""
        durum = getattr(self, durum_attr)
        cerceve = getattr(self, cerceve_attr)
        btn = getattr(self, btn_attr)
        if durum:
            cerceve.pack_forget()
            btn.configure(text="     └  Alt Birimler  ▸")
        else:
            cerceve.pack(fill="x", padx=18, pady=2)
            btn.configure(text="     └  Alt Birimler  ▾")
        setattr(self, durum_attr, not durum)

    def _sidebar_diger_alt_doldur(self):
        """Eski metod — artık _dep_btn_ekle kullanılıyor."""
        pass

    def _toggle_diger_alt(self):
        """Eski metod — artık _toggle_dep kullanılıyor."""
        pass

    def _ayrac(self):
        ctk.CTkFrame(
            self.sidebar, height=1, fg_color="#2D5A8E"
        ).pack(fill="x", padx=12, pady=4)

    def _grup_basligi(self, metin):
        ctk.CTkLabel(
            self.sidebar, text=metin,
            font=("Arial", 9, "bold"),
            text_color="#7AAED0",
            anchor="w"
        ).pack(anchor="w", padx=14, pady=(4, 2))

    def _menu_btn(self, etiket, renk, komut):
        ctk.CTkButton(
            self.sidebar, text=etiket,
            fg_color=renk, hover_color="#333",
            font=("Arial", 11), height=30,
            anchor="w", corner_radius=6,
            command=komut
        ).pack(pady=2, padx=14, fill="x")

    def _dosya_satiri(self, ikon, etiket, altyazi, renk, komut, attr):
        cerceve = ctk.CTkFrame(
            self.sidebar, fg_color=C["card"],
            corner_radius=8
        )
        cerceve.pack(fill="x", padx=18, pady=4)

        ust = ctk.CTkFrame(cerceve, fg_color="transparent")
        ust.pack(fill="x", padx=10, pady=(8, 2))

        ctk.CTkLabel(ust, text=ikon,
                     font=("Arial", 18), width=28
                     ).pack(side="left")
        ctk.CTkLabel(ust, text=etiket,
                     font=("Arial", 12, "bold"),
                     text_color=C["metin"]
                     ).pack(side="left", padx=6)
        ctk.CTkButton(ust, text="Yükle",
                      fg_color=renk, hover_color="#555",
                      font=("Arial", 11), width=60, height=26,
                      corner_radius=6, command=komut
                      ).pack(side="right")

        lbl = ctk.CTkLabel(cerceve, text=f"  {altyazi}  —  yüklenmedi",
                           font=("Arial", 10),
                           text_color=C["metin_koyu"],
                           anchor="w")
        lbl.pack(fill="x", padx=10, pady=(0, 8))
        setattr(self, attr, lbl)

    def _dosya_popup(self):
        """Dosya yükleme popup'ı — 2 dosya, ad değil içerik bazlı."""
        popup = ctk.CTkToplevel(self)
        popup.title("Dosya Yükleme")
        popup.geometry("520x300")
        popup.resizable(False, False)
        popup.configure(fg_color=C["bg"])
        popup.grab_set()
        popup.lift()
        popup.focus_force()

        # Başlık
        bb = ctk.CTkFrame(popup, fg_color="#1A3A5C", corner_radius=0, height=46)
        bb.pack(fill="x")
        bb.pack_propagate(False)
        ctk.CTkLabel(bb, text="  📂  Dosya Yükleme",
                     font=("Arial", 13, "bold"),
                     text_color="#FFFFFF").pack(side="left", padx=14)
        ctk.CTkButton(bb, text="✕", width=32, height=32,
                      fg_color="transparent", hover_color="#555",
                      text_color="#FFFFFF", font=("Arial", 13, "bold"),
                      command=popup.destroy).pack(side="right", padx=8, pady=7)

        ctk.CTkLabel(popup,
                     text="  Dosya adı önemli değil — içerik otomatik tanınır.",
                     font=("Arial", 10), text_color=C["metin_dim"]
                     ).pack(anchor="w", padx=16, pady=(10, 6))

        # 2 kart
        for ikon, ad, aciklama, renk, komut, durum_attr in [
            ("🏭", "Dinamik Rapor",  "Fabrika verisi  —  Kimya / Dow / Diğer",
             C["kimya"], self._yukle_dinamik, "dinamik"),
            ("⚓", "Pregate Sunum",  "Terminal verisi  —  Poliport / Antrepo",
             C["dow"],   self._yukle_pregate, "pregate"),
        ]:
            kart = ctk.CTkFrame(popup, fg_color=C["card"],
                                corner_radius=10,
                                border_width=1, border_color=C["border"])
            kart.pack(fill="x", padx=16, pady=5)
            kart.grid_columnconfigure(1, weight=1)

            ctk.CTkLabel(kart, text=ikon, font=("Arial", 22), width=46
                         ).grid(row=0, column=0, rowspan=2,
                                padx=(12,4), pady=10)
            ctk.CTkLabel(kart, text=ad, font=("Arial", 12, "bold"),
                         text_color=C["metin"], anchor="w"
                         ).grid(row=0, column=1, sticky="w", padx=4, pady=(10,0))
            ctk.CTkLabel(kart, text=aciklama, font=("Arial", 9),
                         text_color=C["metin_dim"], anchor="w"
                         ).grid(row=1, column=1, sticky="w", padx=4, pady=(0,8))

            # Durum etiketi
            durum = getattr(self, f"_durum_{durum_attr}", "  — yüklenmedi")
            renk_lbl = C["yesil"] if "✔" in durum else C["metin_koyu"]
            ctk.CTkLabel(kart, text=durum, font=("Arial", 9),
                         text_color=renk_lbl, anchor="w"
                         ).grid(row=2, column=1, sticky="w", padx=4, pady=(0,6))

            def _yukle(k=komut, p=popup):
                p.destroy()
                self.after(50, lambda: (k(), self._dosya_popup()))

            ctk.CTkButton(kart, text="📂  Yükle",
                          fg_color=renk, hover_color="#444",
                          font=("Arial", 11), width=90, height=30,
                          corner_radius=7, command=_yukle
                          ).grid(row=0, column=2, rowspan=2,
                                 padx=(8,12), pady=10)

        ctk.CTkButton(popup, text="✓  Tamam",
                      fg_color=C["turuncu"], hover_color="#B7600D",
                      font=("Arial", 12, "bold"), height=34,
                      command=popup.destroy
                      ).pack(fill="x", padx=16, pady=(6,12))

    def _dosya_ozet_guncelle(self):
        """Sidebar tek satır durum günceller."""
        isaretler = []
        if not self.motor.df_dinamik.empty: isaretler.append("🏭")
        if not self.motor.df_pregate.empty: isaretler.append("⚓")
        if hasattr(self, "_lbl_dosya_ozet"):
            if isaretler:
                self._lbl_dosya_ozet.configure(
                    text=f"  {'  '.join(isaretler)}  yüklendi",
                    text_color=C["yesil"])
            else:
                self._lbl_dosya_ozet.configure(
                    text="  Henüz dosya yüklenmedi",
                    text_color=C["metin_koyu"])

    def _toggle_alt_birim(self):
        if self._alt_birim_durum:
            self._frame_alt_birim.pack_forget()
            self._btn_alt_ac.configure(text="     └  Alt Birimler  ▸")
        else:
            self._frame_alt_birim.pack(after=self._btn_alt_ac,
                                       fill="x", padx=18, pady=2)
            self._btn_alt_ac.configure(text="     └  Alt Birimler  ▾")
        self._alt_birim_durum = not self._alt_birim_durum

    # =========================================================================
    # ANA ALAN
    # =========================================================================
    def _ana_alan_kur(self):
        self.ana_alan = ctk.CTkFrame(self, fg_color=C["bg"])
        self.ana_alan.grid(row=0, column=1, sticky="nsew", padx=(0,0), pady=0)
        self.ana_alan.grid_rowconfigure(2, weight=1)
        self.ana_alan.grid_columnconfigure(0, weight=1)

        # ── Üst bant ─────────────────────────────────────────────────────────
        ust_bant = ctk.CTkFrame(self.ana_alan, fg_color=C["card"],
                                corner_radius=0, height=48)
        ust_bant.grid(row=0, column=0, sticky="ew", pady=(0, 0))
        ust_bant.grid_columnconfigure(0, weight=1)
        ust_bant.grid_propagate(False)

        self.lbl_sayfa_basligi = ctk.CTkLabel(
            ust_bant,
            text=f"Hoş Geldiniz  —  {UYGULAMA_ADI}",
            font=("Arial", 15, "bold"),
            text_color=C["altin"],
            anchor="w"
        )
        self.lbl_sayfa_basligi.grid(row=0, column=0, sticky="w",
                                    padx=16, pady=12)

        # Sağ taraf: imza + tarih + tema + indir
        sag = ctk.CTkFrame(ust_bant, fg_color="transparent")
        sag.grid(row=0, column=1, sticky="e", padx=14)

        # S. SEYMEN imzası — her sayfada görünür
        ctk.CTkLabel(
            sag,
            text=f"© {HAZIRLAYEN}",
            font=("Arial", 11, "italic"),
            text_color=C["altin_dim"]
        ).pack(side="left", padx=(0, 14))

        self.lbl_tarih = ctk.CTkLabel(
            sag,
            text=datetime.datetime.now().strftime("%d.%m.%Y  %H:%M"),
            font=("Arial", 12),
            text_color=C["metin_dim"],
        )
        self.lbl_tarih.pack(side="left", padx=(0, 10))

        # Tema değiştirici
        self._tema_var = ctk.StringVar(value="Dark")
        self.btn_tema = ctk.CTkButton(
            sag,
            text="☀  Açık Tema",
            width=110, height=30,
            fg_color=C["card2"],
            hover_color=C["hover"],
            border_width=1,
            border_color=C["border"],
            text_color=C["metin_dim"],
            font=("Arial", 11),
            command=self._tema_degistir
        )
        self.btn_tema.pack(side="left", padx=4)

        # Tümünü İndir butonu
        self.btn_tumunu_indir = ctk.CTkButton(
            sag,
            text="⬇  Tümünü İndir",
            width=130, height=30,
            fg_color=C["altin_dim"],
            hover_color="#6B5500",
            text_color=C["altin"],
            font=("Arial", 11, "bold"),
            command=self._export_excel
        )
        self.btn_tumunu_indir.pack(side="left", padx=4)

        # ── KPI bandı — kompakt ──────────────────────────────────────────────
        kpi = ctk.CTkFrame(self.ana_alan, fg_color="transparent")
        kpi.grid(row=1, column=0, sticky="ew", padx=10, pady=(4, 2))

        self.kpi_toplam = self._kpi_kart(kpi, "Toplam Araç",    "—", C["mavi"],       0)
        self.kpi_onay   = self._kpi_kart(kpi, "Onaylanan",      "—", C["yesil"],      1)
        self.kpi_red    = self._kpi_kart(kpi, "Reddedilen",     "—", C["kirmizi"],    2)
        self.kpi_oran   = self._kpi_kart(kpi, "Onay Oranı",     "—", C["camgobegi"],  3)

        # ── İçerik alanı ─────────────────────────────────────────────────────
        self.icerik = ctk.CTkFrame(self.ana_alan, fg_color=C["bg"])
        self.icerik.grid(row=2, column=0, sticky="nsew", padx=0, pady=0)
        self.icerik.grid_rowconfigure(1, weight=1)
        self.icerik.grid_columnconfigure(0, weight=1)

        # Nakliyeci alt menü (başta gizli)
        self._nakliye_menu = ctk.CTkFrame(self.icerik, fg_color=C["card"],
                                          corner_radius=8)
        for txt, renk, fn in [
            ("📋  Tüm Kayıtlar",          "#34495E", self._nakliye_tum),
            ("✅  Onaylananlar",           "#1E8449", self._nakliye_onay),
            ("❌  Reddilenler & Matrix",   "#922B21", self._nakliye_red),
        ]:
            ctk.CTkButton(
                self._nakliye_menu, text=txt,
                fg_color=renk, hover_color="#555",
                font=("Arial", 12), height=36,
                command=fn
            ).pack(side="left", padx=10, pady=8, expand=True, fill="x")

        # Metin rapor kutusu
        self.txt_rapor = ctk.CTkTextbox(
            self.icerik,
            font=("Consolas", 13),
            fg_color=C["card"],
            text_color=C["metin"],
            corner_radius=10,
            border_width=1,
            border_color=C["border"]
        )
        self._metin_tagleri_ayarla()
        # Metin kutusu sağ tık menüsü
        self._metin_menu = self._metin_menu_olustur()
        tb = self.txt_rapor._textbox
        tb.bind("<Button-3>", self._metin_sag_tik)
        tb.bind("<Button-2>", self._metin_sag_tik)  # macOS

        # Seçme + kopyalama (Mac & Windows)
        def _kopyala(e):
            try:
                sec = tb.get("sel.first", "sel.last")
                self.clipboard_clear()
                self.clipboard_append(sec)
            except Exception:
                pass
            return "break"
        def _tumsec(e):
            tb.tag_add("sel", "1.0", "end")
            return "break"
        tb.bind("<Control-c>", _kopyala)
        tb.bind("<Command-c>", _kopyala)
        tb.bind("<Control-a>", _tumsec)
        tb.bind("<Command-a>", _tumsec)
        # Yazmayı engelle ama seçime izin ver
        def _engelle(e):
            izinli = {"Up","Down","Left","Right","Home","End","Prior","Next",
                      "Shift_L","Shift_R","Control_L","Control_R",
                      "Command_L","Command_R","Meta_L","Meta_R"}
            if e.keysym in izinli:
                return
            if (e.state & 0x4 or e.state & 0x8 or e.state & 0x100) and \
               e.keysym.lower() in ("c","a"):
                return
            return "break"
        tb.bind("<Key>", _engelle)

        # Dashboard çerçevesi
        self.dash = ctk.CTkScrollableFrame(
            self.icerik, fg_color=C["bg"],
            scrollbar_button_color=C["border"],
            scrollbar_button_hover_color=C["hover"]
        )

        # Tablo çerçevesi
        self.tablo_cerceve = ctk.CTkFrame(
            self.icerik, fg_color=C["card"],
            corner_radius=10
        )
        self.tablo_cerceve.grid_rowconfigure(0, weight=1)
        self.tablo_cerceve.grid_columnconfigure(0, weight=1)

        sy = ttk.Scrollbar(self.tablo_cerceve, orient="vertical",
                           style="SR.Scrollbar")
        sy.grid(row=0, column=1, sticky="ns")
        sx = ttk.Scrollbar(self.tablo_cerceve, orient="horizontal",
                           style="SR.Scrollbar")
        sx.grid(row=1, column=0, sticky="ew")
        self.tablo = ttk.Treeview(
            self.tablo_cerceve,
            style="SR.Treeview",
            yscrollcommand=sy.set,
            xscrollcommand=sx.set
        )
        self.tablo.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        sy.configure(command=self.tablo.yview)
        sx.configure(command=self.tablo.xview)

        # Tablo sağ tık menüsü
        self._tablo_menu = self._kontekst_menu_olustur()
        self.tablo.bind("<Button-3>",  self._tablo_sag_tik)
        self.tablo.bind("<Button-2>",  self._tablo_sag_tik)  # macOS

    # =========================================================================
    # SAĞ TIK MENÜSÜ  —  Kopyala / Tümünü Seç / Excel Kopyası
    # =========================================================================
    def _kontekst_menu_olustur(self):
        """Tablo için sağ tık context menüsü oluşturur."""
        from tkinter import Menu
        menu = Menu(self, tearoff=0,
                    bg="#1A2438", fg=C["metin"],
                    activebackground="#2E86DE",
                    activeforeground="#FFFFFF",
                    font=("Arial", 12),
                    bd=0, relief="flat")
        menu.add_command(label="  📋  Seçili Satırı Kopyala",
                         command=self._tablo_secili_kopyala)
        menu.add_command(label="  📄  Tüm Tabloyu Kopyala",
                         command=self._tablo_tumu_kopyala)
        menu.add_separator()
        menu.add_command(label="  ✅  Tüm Satırları Seç",
                         command=self._tablo_tumunu_sec)
        menu.add_separator()
        menu.add_command(label="  💾  Excel'e Aktar (Bu Sayfa)",
                         command=self._tablo_excele_aktar)
        return menu

    def _metin_menu_olustur(self):
        """Metin kutusu için sağ tık menüsü."""
        from tkinter import Menu
        menu = Menu(self, tearoff=0,
                    bg="#1A2438", fg=C["metin"],
                    activebackground="#2E86DE",
                    activeforeground="#FFFFFF",
                    font=("Arial", 12),
                    bd=0, relief="flat")
        menu.add_command(label="  📋  Seçili Metni Kopyala",
                         command=self._metin_secili_kopyala)
        menu.add_command(label="  📄  Tüm Metni Kopyala",
                         command=self._metin_tumu_kopyala)
        menu.add_separator()
        menu.add_command(label="  ✅  Tümünü Seç",
                         command=self._metin_tumunu_sec)
        return menu

    def _tablo_sag_tik(self, event):
        """Tabloda sağ tıkta satırı seç ve menü aç."""
        item = self.tablo.identify_row(event.y)
        if item:
            self.tablo.selection_set(item)
        try:
            self._tablo_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._tablo_menu.grab_release()

    def _metin_sag_tik(self, event):
        try:
            self._metin_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self._metin_menu.grab_release()

    def _genel_tablo_kopyala(self, tablo_widget, tumu=False):
        """
        Herhangi bir Treeview widget'ından seçili veya tüm satırları TSV olarak kopyalar.
        tumu=True  → tüm satırlar
        tumu=False → sadece seçili satırlar
        """
        try:
            basliklar = [tablo_widget.heading(c)["text"]
                         for c in tablo_widget["columns"]]
            satirlar = ["\t".join(basliklar)]
            if tumu:
                items = tablo_widget.get_children()
            else:
                items = tablo_widget.selection()
            if not items:
                items = tablo_widget.get_children()
            for item in items:
                degerler = [str(v) for v in tablo_widget.item(item, "values")]
                satirlar.append("\t".join(degerler))
            self.clipboard_clear()
            self.clipboard_append("\n".join(satirlar))
            self._bildirim(f"✔  {len(satirlar)-1} satır panoya kopyalandı.")
        except Exception as e:
            self._bildirim(f"⚠  Kopyalama hatası: {e}")

    def _tablo_secili_kopyala(self):
        """Seçili tablo satırlarını TSV olarak panoya kopyalar."""
        secili = self.tablo.selection()
        if not secili:
            return
        satirlar = []
        # Başlık satırı
        basliklar = [self.tablo.heading(c)["text"]
                     for c in self.tablo["columns"]]
        satirlar.append("\t".join(basliklar))
        for item in secili:
            degerler = [str(v) for v in self.tablo.item(item, "values")]
            satirlar.append("\t".join(degerler))
        self.clipboard_clear()
        self.clipboard_append("\n".join(satirlar))
        self._bildirim("✔  Seçili satır(lar) panoya kopyalandı.")

    def _tablo_tumu_kopyala(self):
        """Tablodaki tüm satırları TSV olarak panoya kopyalar."""
        basliklar = [self.tablo.heading(c)["text"]
                     for c in self.tablo["columns"]]
        satirlar = ["\t".join(basliklar)]
        for item in self.tablo.get_children():
            degerler = [str(v) for v in self.tablo.item(item, "values")]
            satirlar.append("\t".join(degerler))
        self.clipboard_clear()
        self.clipboard_append("\n".join(satirlar))
        self._bildirim(f"✔  {len(satirlar)-1} satır panoya kopyalandı.")

    def _tablo_tumunu_sec(self):
        for item in self.tablo.get_children():
            self.tablo.selection_add(item)

    def _tablo_excele_aktar(self):
        """Ekrandaki tabloyu tek sayfalık Excel olarak kaydeder."""
        basliklar = [self.tablo.heading(c)["text"]
                     for c in self.tablo["columns"]]
        satirlar = []
        for item in self.tablo.get_children():
            satirlar.append(list(self.tablo.item(item, "values")))
        if not satirlar:
            messagebox.showwarning("Uyarı", "Tabloda veri yok.")
            return
        df = pd.DataFrame(satirlar, columns=basliklar)
        yol = filedialog.asksaveasfilename(
            title="Tabloyu Kaydet",
            defaultextension=".xlsx",
            initialfile=f"Tablo_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not yol:
            return
        try:
            df.to_excel(yol, index=False)
            self._bildirim(f"✔  Excel kaydedildi: {os.path.basename(yol)}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")

    def _metin_secili_kopyala(self):
        try:
            metin = self.txt_rapor._textbox.get("sel.first", "sel.last")
            self.clipboard_clear()
            self.clipboard_append(metin)
            self._bildirim("✔  Seçili metin kopyalandı.")
        except Exception:
            self._bildirim("⚠  Kopyalanacak seçili metin yok.")

    def _metin_tumu_kopyala(self):
        metin = self.txt_rapor.get("1.0", "end")
        self.clipboard_clear()
        self.clipboard_append(metin)
        self._bildirim("✔  Tüm rapor metni kopyalandı.")

    def _metin_tumunu_sec(self):
        self.txt_rapor._textbox.tag_add("sel", "1.0", "end")

    def _bildirim(self, mesaj):
        """Başlık çubuğunda kısa bildirim gösterir."""
        onceki = self.title()
        self.title(f"  {mesaj}   —   {UYGULAMA_ADI}")
        self.after(2500, lambda: self.title(onceki))

    def _kpi_kart(self, parent, baslik, deger, renk, idx):
        kart = ctk.CTkFrame(parent, fg_color=C["card"],
                            corner_radius=8,
                            border_width=1, border_color=C["border"])
        kart.grid(row=0, column=idx, padx=4, sticky="nsew")
        parent.grid_columnconfigure(idx, weight=1)

        ctk.CTkLabel(kart, text=baslik,
                     font=("Arial", 10), text_color=C["metin_dim"]
                     ).pack(pady=(8, 1))
        lbl = ctk.CTkLabel(kart, text=deger,
                           font=("Arial", 22, "bold"), text_color=renk)
        lbl.pack(pady=(1, 2))

        # Reddedilen kartı için alt kırılım frame
        if idx == 2:
            self._kpi_red_alt_frame = ctk.CTkFrame(kart, fg_color="transparent")
            self._kpi_red_alt_frame.pack(pady=(0, 6))
        else:
            ctk.CTkFrame(kart, height=6, fg_color="transparent").pack()

        return lbl

    def _metin_tagleri_ayarla(self):
        t = self.txt_rapor._textbox
        # Temaya bağlı renkler
        h1_bg  = C["tag_h1_bg"]
        tek_bg = C["tag_tek_bg"]
        cift_bg = C["tag_cift_bg"]
        metin_renk = C["metin"]

        t.tag_configure("h1",
                        background=h1_bg, foreground=C["altin"],
                        font=("Arial", 15, "bold"),
                        spacing1=16, spacing3=16)
        t.tag_configure("h2",
                        foreground=C["mavi"],
                        font=("Arial", 13, "bold"),
                        spacing1=10, spacing3=4)
        t.tag_configure("tek",
                        background=tek_bg, foreground=metin_renk,
                        font=("Consolas", 12),
                        spacing1=5, spacing3=5,
                        tabs=("460","620","780"))
        t.tag_configure("cift",
                        background=cift_bg, foreground=metin_renk,
                        font=("Consolas", 12),
                        spacing1=5, spacing3=5,
                        tabs=("460","620","780"))
        t.tag_configure("uyari",
                        foreground=C["turuncu"],
                        font=("Arial", 12, "bold"),
                        spacing1=6, spacing3=6)
        t.tag_configure("basarili",
                        foreground=C["yesil"],
                        font=("Arial", 12),
                        spacing1=4)
        t.tag_configure("eksik",
                        foreground=C["kirmizi"],
                        font=("Consolas", 12),
                        spacing1=5, spacing3=5)
        # Metin kutusu arka planı temaya uysun
        self.txt_rapor.configure(fg_color=C["card"], text_color=metin_renk)

    # =========================================================================
    # TEMA DEĞİŞTİRİCİ
    # =========================================================================
    def _tema_degistir(self):
        global C, TEMA_MOD
        if TEMA_MOD == "Dark":
            TEMA_MOD = "Light"
            ctk.set_appearance_mode("Light")
            C = _palet_hesapla("Light")
            self.btn_tema.configure(text="🌙  Koyu Tema",
                                    fg_color=C["card2"],
                                    border_color=C["border"],
                                    text_color=C["metin_dim"])
        else:
            TEMA_MOD = "Dark"
            ctk.set_appearance_mode("Dark")
            C = _palet_hesapla("Dark")
            self.btn_tema.configure(text="☀  Açık Tema",
                                    fg_color=C["card2"],
                                    border_color=C["border"],
                                    text_color=C["metin_dim"])
        # Metin tag renklerini güncelle
        self._metin_tagleri_ayarla()
        # Sayfayı yenile
        self._ana_sayfa()

    # =========================================================================
    # DEPARTMAN BAZLI TEK SAYFA EXCEL
    # =========================================================================
    def _dept_excel(self, departman):
        """Tek departmanın verisini ayrı bir Excel dosyasına kaydeder."""
        if departman in ("Kimya","Dow","Diğer"):
            df = self.motor._dinamik_filtrele(departman)
            kaynak = "Dinamik"
        elif departman == "Poliport Terminal":
            df = self.motor._pregate_filtrele("Poliport Terminal")
            kaynak = "Terminal"
        elif departman == "Antrepo":
            df = self.motor._pregate_filtrele("Antrepo")
            kaynak = "Antrepo"
        else:
            df = pd.DataFrame()
            kaynak = departman

        if df.empty:
            messagebox.showwarning("Uyarı", f"{departman} verisi bulunamadı.")
            return

        yol = filedialog.asksaveasfilename(
            title=f"{departman} Verisini Kaydet",
            defaultextension=".xlsx",
            initialfile=f"{kaynak}_{departman.replace(' ','_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not yol:
            return
        try:
            with pd.ExcelWriter(yol, engine="openpyxl") as w:
                df.to_excel(w, sheet_name=departman[:31], index=False)
                # Özet satır
                t, o, r = (self.motor.dinamik_kpi(df)
                           if kaynak == "Dinamik"
                           else self.motor.pregate_kpi(df))
                ozet = pd.DataFrame([{
                    "Departman": departman,
                    "Toplam": t, "Onaylanan": o, "Reddedilen": r,
                    "Onay Oranı": f"%{round(o/t*100,1)}" if t > 0 else "—"
                }])
                ozet.to_excel(w, sheet_name="Özet", index=False)
            self._bildirim(f"✔  {departman} Excel kaydedildi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")
    def _ekran_hazirla(self, baslik, mod="metin", menu=False):
        self.lbl_sayfa_basligi.configure(text=baslik)
        self.txt_rapor.delete("1.0", "end")

        # Hepsini gizle
        self._nakliye_menu.grid_forget()
        self.txt_rapor.grid_forget()
        self.dash.grid_forget()
        self.tablo_cerceve.grid_forget()

        # Önceki özel frame temizle
        if hasattr(self, "_ozel_frame") and self._ozel_frame:
            try:
                self._ozel_frame.destroy()
            except Exception:
                pass
            self._ozel_frame = None

        satir = 0
        if menu:
            self._nakliye_menu.grid(row=0, column=0, sticky="ew",
                                    pady=(0, 10))
            satir = 1

        self.icerik.grid_rowconfigure(satir, weight=1)

        if mod == "dashboard":
            self.dash.grid(row=satir, column=0, sticky="nsew")
        elif mod == "tablo":
            self.tablo_cerceve.grid(row=satir, column=0, sticky="nsew")
        elif mod == "ozel":
            self._ozel_frame = ctk.CTkFrame(
                self.icerik, fg_color="transparent")
            self._ozel_frame.grid(row=satir, column=0, sticky="nsew")
            self._ozel_frame.grid_rowconfigure(0, weight=1)
            self._ozel_frame.grid_columnconfigure(0, weight=1)
        else:
            self.txt_rapor.grid(row=satir, column=0, sticky="nsew")

    def _yaz(self, metin, stil=None):
        if stil:
            self.txt_rapor._textbox.insert("end", metin, stil)
        else:
            self.txt_rapor.insert("end", metin)

    def _fmt(self, isim, v1, v2=None, v3=None, s="▪", eksik=False):
        ism = str(isim).upper().strip()[:50]
        ek  = "  ⚠" if eksik else ""
        if v2 is not None and v3 is not None:
            return f" {s} {ism}\tToplam: {v1}\t| Onay: {v2}\t| Red: {v3}\n"
        return f" {s} {ism}{ek}\t: {v1}\n"

    def _kpi_guncelle(self, toplam, onay, red, red_detay=None):
        self.kpi_toplam.configure(text=str(toplam))
        self.kpi_onay.configure(text=str(onay))
        self.kpi_red.configure(text=str(red))
        oran = f"%{round(onay/toplam*100,1)}" if toplam > 0 else "—"
        self.kpi_oran.configure(text=oran)

        # Reddedilen kartı alt kırılım — Pregate / Sevkiyat / Operasyon
        if hasattr(self, "_kpi_red_alt_frame"):
            for w in self._kpi_red_alt_frame.winfo_children():
                w.destroy()
            if red_detay:
                for tip, sayi, renk in [
                    ("Pregate",   red_detay.get("Pregate Red",   0), "#E74C3C"),
                    ("Sevkiyat",  red_detay.get("Sevkiyat Red",  0), "#E67E22"),
                    ("Operasyon", red_detay.get("Operasyon Red", 0), "#2E86DE"),
                ]:
                    ctk.CTkLabel(
                        self._kpi_red_alt_frame,
                        text=f"{tip}: {sayi}",
                        font=("Arial", 10, "bold"),
                        text_color=renk
                    ).pack(side="left", padx=6)

    # =========================================================================
    # ARAÇ / PLAKA ARAMA SAYFASI
    # =========================================================================
    def _goster_arama(self):
        self._ekran_hazirla("🔍  Araç / Plaka / Nakliyeci Arama", mod="ozel")

        # Ana çerçeve — _ozel_frame içine yerleştir
        ana = ctk.CTkFrame(self._ozel_frame, fg_color="transparent")
        ana.grid(row=0, column=0, sticky="nsew")
        ana.grid_rowconfigure(1, weight=1)
        ana.grid_columnconfigure(0, weight=1)
        self._arama_ana = ana

        # ── Filtre bandı ──────────────────────────────────────────────────────
        filtre_kart = ctk.CTkFrame(ana, fg_color=C["card"],
                                   corner_radius=10,
                                   border_width=1, border_color=C["border"])
        filtre_kart.grid(row=0, column=0, sticky="ew", pady=(0, 10))
        filtre_kart.grid_columnconfigure((0,1,2,3,4,5), weight=1)

        # Satır 0 — etiketler
        for i, lbl in enumerate(["Plaka / Dorse / Tank / Nakliyeci",
                                  "Durum Filtresi",
                                  "Tarih Başlangıç", "Tarih Bitiş",
                                  "", ""]):
            ctk.CTkLabel(filtre_kart, text=lbl,
                         font=("Arial", 11), text_color=C["metin_dim"]
                         ).grid(row=0, column=i, padx=10, pady=(10,2), sticky="w")

        # Satır 1 — giriş alanları
        self._arama_entry = ctk.CTkEntry(
            filtre_kart,
            placeholder_text="Örn: 34AB1234  veya  NAKLIYECI ADI",
            font=("Arial", 13), height=36, width=280
        )
        self._arama_entry.grid(row=1, column=0, padx=10, pady=(0,10), sticky="ew")
        self._arama_entry.bind("<Return>", lambda e: self._arama_calistir())

        self._arama_durum = ctk.CTkOptionMenu(
            filtre_kart,
            values=["Tümü","Onay","Red"],
            font=("Arial", 12), height=36,
            fg_color=C["card2"], button_color=C["border"],
            dropdown_fg_color=C["card"]
        )
        self._arama_durum.grid(row=1, column=1, padx=10, pady=(0,10), sticky="ew")
        self._arama_durum.set("Tümü")

        self._arama_tarih_bas = ctk.CTkEntry(
            filtre_kart, placeholder_text="GG.AA.YYYY",
            font=("Arial", 12), height=36
        )
        self._arama_tarih_bas.grid(row=1, column=2, padx=10, pady=(0,10), sticky="ew")

        self._arama_tarih_bit = ctk.CTkEntry(
            filtre_kart, placeholder_text="GG.AA.YYYY",
            font=("Arial", 12), height=36
        )
        self._arama_tarih_bit.grid(row=1, column=3, padx=10, pady=(0,10), sticky="ew")

        ctk.CTkButton(
            filtre_kart, text="🔍  Ara",
            fg_color=C["mavi"], hover_color="#1460A0",
            font=("Arial", 13, "bold"), height=36,
            command=self._arama_calistir
        ).grid(row=1, column=4, padx=10, pady=(0,10), sticky="ew")

        ctk.CTkButton(
            filtre_kart, text="✖  Temizle",
            fg_color=C["card2"], hover_color=C["border"],
            text_color=C["metin_dim"],
            font=("Arial", 12), height=36,
            command=self._arama_temizle
        ).grid(row=1, column=5, padx=10, pady=(0,10), sticky="ew")

        # ── Sonuç alanı ───────────────────────────────────────────────────────
        sonuc_cerceve = ctk.CTkFrame(ana, fg_color="transparent")
        sonuc_cerceve.grid(row=1, column=0, sticky="nsew")
        sonuc_cerceve.grid_rowconfigure(0, weight=1)
        sonuc_cerceve.grid_columnconfigure(0, weight=1)
        self._arama_sonuc_cerceve = sonuc_cerceve

        # Başlangıç mesajı
        self._arama_bilgi = ctk.CTkLabel(
            sonuc_cerceve,
            text="Arama yapmak için plaka, dorse no, tank no veya nakliyeci adı girin.\n"
                 "Tüm yüklü dosyalar taranır  —  sonuçlar kaynaklarına göre ayrılır.",
            font=("Arial", 13), text_color=C["metin_dim"],
            justify="center"
        )
        self._arama_bilgi.place(relx=0.5, rely=0.45, anchor="center")

        # Sonuç paneli (başta gizli)
        self._arama_panel = None

    def _arama_temizle(self):
        self._arama_entry.delete(0, "end")
        self._arama_durum.set("Tümü")
        self._arama_tarih_bas.delete(0, "end")
        self._arama_tarih_bit.delete(0, "end")
        if self._arama_panel:
            self._arama_panel.destroy()
            self._arama_panel = None
        if hasattr(self, "_arama_bilgi"):
            self._arama_bilgi.place(relx=0.5, rely=0.45, anchor="center")

    def _tarih_parse(self, metin):
        """GG.AA.YYYY → YYYY-MM-DD çevirir, hata olursa "" döner."""
        metin = metin.strip()
        if not metin:
            return ""
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.datetime.strptime(metin, fmt).strftime("%Y-%m-%d")
            except ValueError:
                continue
        return ""

    def _arama_calistir(self):
        sorgu = self._arama_entry.get().strip()
        if not sorgu:
            messagebox.showwarning("Uyarı", "Lütfen arama terimi girin.")
            return

        filtreler = {
            "durum":     self._arama_durum.get(),
            "tarih_bas": self._tarih_parse(self._arama_tarih_bas.get()),
            "tarih_bit": self._tarih_parse(self._arama_tarih_bit.get()),
        }

        # Önceki sonuç panelini temizle
        if self._arama_panel:
            self._arama_panel.destroy()
        if hasattr(self, "_arama_bilgi"):
            self._arama_bilgi.place_forget()

        # Arama yap
        sonuclar = self.motor.arac_ara(sorgu.upper(), filtreler)

        toplam = sum(len(v) for v in sonuclar.values())

        # Sonuç paneli
        panel = ctk.CTkScrollableFrame(
            self._arama_sonuc_cerceve,
            fg_color="transparent",
            scrollbar_button_color=C["border"]
        )
        panel.grid(row=0, column=0, sticky="nsew")
        panel.grid_columnconfigure(0, weight=1)
        self._arama_panel = panel

        if toplam == 0:
            ctk.CTkLabel(
                panel,
                text=f"  '{sorgu}'  için hiçbir dosyada eşleşme bulunamadı.",
                font=("Arial", 13), text_color=C["turuncu"]
            ).pack(pady=30)
            return

        # Özet başlık
        ctk.CTkLabel(
            panel,
            text=f"  '{sorgu}'  için toplam  {toplam}  kayıt bulundu",
            font=("Arial", 14, "bold"), text_color=C["altin"],
            anchor="w"
        ).pack(fill="x", padx=10, pady=(6,2))

        # Her kaynak için ayrı tablo
        kaynak_config = {
            "dinamik": ("🏭  Dinamik Rapor (Fabrika)",    C["kimya"]),
            "pregate": ("⚓  Pregate Sunum (Terminal)",    C["terminal"]),
            "ek":      ("📄  Ek Dosya",                    C["ek"]),
        }

        for anahtar, (baslik, renk) in kaynak_config.items():
            df_r = sonuclar[anahtar]
            if df_r.empty:
                continue

            # Bölüm başlığı + indir butonu
            bolum = ctk.CTkFrame(panel, fg_color=C["card2"], corner_radius=8)
            bolum.pack(fill="x", padx=10, pady=(10, 0))
            bolum.grid_columnconfigure(0, weight=1)

            ctk.CTkLabel(
                bolum,
                text=f"  {baslik}  —  {len(df_r)} kayıt",
                font=("Arial", 13, "bold"), text_color=renk,
                anchor="w"
            ).grid(row=0, column=0, padx=12, pady=8, sticky="w")

            ctk.CTkButton(
                bolum,
                text="📥 Excel İndir",
                width=110, height=26,
                fg_color=renk, hover_color="#333",
                font=("Arial", 11),
                command=lambda d=df_r, k=anahtar, s=sorgu:
                    self._arama_excele_kaydet(d, k, s)
            ).grid(row=0, column=1, padx=12, pady=6, sticky="e")

            # Tablo
            tablo_cerceve = ctk.CTkFrame(panel, fg_color=C["card"],
                                          corner_radius=8)
            tablo_cerceve.pack(fill="x", padx=10, pady=(0,4))
            tablo_cerceve.grid_rowconfigure(0, weight=1)
            tablo_cerceve.grid_columnconfigure(0, weight=1)

            sy = ttk.Scrollbar(tablo_cerceve, orient="vertical")
            sy.grid(row=0, column=1, sticky="ns")
            sx = ttk.Scrollbar(tablo_cerceve, orient="horizontal")
            sx.grid(row=1, column=0, sticky="ew")

            tablo = ttk.Treeview(
                tablo_cerceve,
                style="SR.Treeview",
                yscrollcommand=sy.set,
                xscrollcommand=sx.set,
                height=min(len(df_r), 12)
            )
            tablo.grid(row=0, column=0, sticky="nsew")
            sy.configure(command=tablo.yview)
            sx.configure(command=tablo.xview)

            # Sütunlar — önce önemli sütunlar
            oncelik = ["Eşleşme Alanı","Kaynak"]
            geri_kalan = [c for c in df_r.columns if c not in oncelik]
            cols = oncelik + geri_kalan

            tablo["columns"] = cols
            tablo.column("#0", width=0, stretch="no")
            tablo.heading("#0", text="")

            for col in cols:
                if col in ("Kaynak","Eşleşme Alanı"):
                    w = 130
                else:
                    try:
                        max_ic = df_r[col].astype(str).str.len().max()
                        w = max(min(int(max_ic) * 8 + 20, 200), 80)
                    except Exception:
                        w = 120
                tablo.column(col, anchor="w", width=w, minwidth=60)
                tablo.heading(col, text=col, anchor="w")

            for i, (_, row) in enumerate(df_r.iterrows()):
                vals = [str(row[c]) if str(row[c]) != "nan" else ""
                        for c in cols]
                tag = "tek_r" if i % 2 == 0 else "cift_r"
                tablo.insert("", "end", values=vals, tags=(tag,))

            tablo.tag_configure("tek_r",  background="#141C2B",
                                font=("Consolas",10))
            tablo.tag_configure("cift_r", background="#111827",
                                font=("Consolas",10))

            # Sağ tık kopyala
            menu_ref = self._kontekst_menu_olustur()
            def _sag_tik(e, t=tablo, m=menu_ref):
                item = t.identify_row(e.y)
                if item: t.selection_set(item)
                try: m.tk_popup(e.x_root, e.y_root)
                finally: m.grab_release()
            tablo.bind("<Button-3>", _sag_tik)
            tablo.bind("<Button-2>", _sag_tik)

    def _arama_excele_kaydet(self, df, kaynak, sorgu):
        """Arama sonucu tek kaynağı Excel olarak kaydeder."""
        yol = filedialog.asksaveasfilename(
            title=f"{kaynak} Arama Sonucunu Kaydet",
            defaultextension=".xlsx",
            initialfile=f"Arama_{sorgu.replace(' ','_')}_{kaynak}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not yol:
            return
        try:
            df.to_excel(yol, index=False)
            self._bildirim(f"✔  {kaynak} arama sonucu kaydedildi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")

    # =========================================================================
    # HOŞ GELDİN EKRANI
    # =========================================================================
    def _hosgeldin_goster(self):
        self._ekran_hazirla(f"Hoş Geldiniz  —  {UYGULAMA_ADI}", mod="metin")
        self._yaz(f"\n  {'─'*60}\n", "uyari")
        self._yaz(f"   {UYGULAMA_ADI}  {UYGULAMA_SURUMU}  —  Pregate Operasyon Yönetim Sistemi\n", "h1")
        self._yaz(f"   Hazırlayan: {HAZIRLAYEN}\n", "h2")
        self._yaz(f"  {'─'*60}\n\n", "uyari")
        self._yaz("  BAŞLAMAK İÇİN:\n", "h2")
        satırlar = [
            ("1", "Sol menüden  'Dinamik Rapor Yükle'  butonuna tıklayın"),
            ("2", "Sol menüden  'Pregate Sunum Yükle'  butonuna tıklayın"),
            ("3", "İsteğe bağlı olarak  'Ek Dosya'  yükleyebilirsiniz"),
            ("4", "Her dosya ayrı ayrı yüklenebilir, hepsi bağımsız çalışır"),
            ("5", "'Ana Sayfa' tuşuna basarak genel özeti görüntüleyin"),
        ]
        for i, (no, txt) in enumerate(satırlar):
            tag = "tek" if int(no) % 2 else "cift"
            self._yaz(f"  {no}.  {txt}\n", tag)

        self._yaz("\n\n  DEPARTMANLAR:\n", "h2")
        deps = [
            ("🏭 Dinamik Rapor", "Kimya  /  Dow  /  Diğer  (Fabrika)"),
            ("⚓ Pregate Sunum", "Poliport Terminal  /  Antrepo"),
            ("📄 Ek Dosya",      "Herhangi bir Excel kaynağı"),
        ]
        for i, (dep, acik) in enumerate(deps):
            tag = "tek" if i % 2 == 0 else "cift"
            self._yaz(f"  {dep:<28}{acik}\n", tag)

        self._yaz("\n\n  ÇIKTILAR:\n", "h2")
        self._yaz("  📊  Excel  →  Tüm sayfalar ve nakliyeci matrix\n", "tek")
        self._yaz("  🎬  PPTX   →  Yönetim sunumu formatında rapor\n", "cift")
        self._yaz("\n")

    # =========================================================================
    # DOSYA YÜKLEME
    # =========================================================================
    def _dosya_sec(self, baslik):
        return filedialog.askopenfilename(
            title=baslik,
            filetypes=[
                ("Excel Dosyaları", "*.xlsx *.xls"),
                ("Tüm Dosyalar", "*.*")
            ]
        )

    def _yukle_dinamik(self):
        yol = self._dosya_sec("Dinamik Rapor Excel dosyasını seçin")
        if not yol:
            return
        try:
            cikarilan = self.motor.yukle_dinamik(yol)
            kisa = os.path.basename(yol)
            n = len(self.motor.df_dinamik)
            self._durum_dinamik = f"  ✔  {kisa}  ({n} satır)"
            tekil_bilgi = (f"\n\n⚠  {cikarilan} tekrarlı kayıt çıkarıldı\n"
                           f"(Plaka + Tarih bazlı tekilleştirme)") if cikarilan > 0 else ""
            messagebox.showinfo(
                "Başarılı",
                f"Dinamik Rapor yüklendi.\n\n"
                f"Dosya:  {kisa}\n"
                f"Toplam satır (tekil):  {n}"
                f"{tekil_bilgi}"
            )
            self._dosya_ozet_guncelle()
            self._sidebar_kur()   # Diğer alt birimleri dolsun
            self._ana_sayfa()
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okunamadı:\n{e}")

    def _yukle_pregate(self):
        yol = self._dosya_sec("Pregate Sunum Excel dosyasını seçin")
        if not yol:
            return
        try:
            cikarilan = self.motor.yukle_pregate(yol)
            kisa = os.path.basename(yol)
            n = len(self.motor.df_pregate)
            self._durum_pregate = f"  ✔  {kisa}  ({n} satır)"
            tekil_bilgi = (f"\n\n⚠  {cikarilan} tekrarlı kayıt çıkarıldı\n"
                           f"(Plaka + Tarih bazlı tekilleştirme)") if cikarilan > 0 else ""
            messagebox.showinfo(
                "Başarılı",
                f"Pregate Sunum yüklendi.\n\n"
                f"Dosya:  {kisa}\n"
                f"Toplam satır (tekil):  {n}"
                f"{tekil_bilgi}"
            )
            self._dosya_ozet_guncelle()
            self._ana_sayfa()
        except Exception as e:
            messagebox.showerror("Hata", f"Dosya okunamadı:\n{e}")

    # =========================================================================
    # ANA SAYFA — DASHBOARD
    # =========================================================================
    def _ana_sayfa(self):
        self._ekran_hazirla("🏠  GENEL OPERASYON DASHBOARD'U", mod="dashboard")
        for w in self.dash.winfo_children():
            w.destroy()
        self.dash.grid_columnconfigure((0, 1, 2), weight=1)

        ozet = self.motor.genel_ozet()
        self._kpi_guncelle(ozet["toplam"], ozet["onay"], ozet["red"])

        satir = 0

        # ── Fabrika bölümü ───────────────────────────────────────────────────
        self._bolum_basligi("🏭  DİNAMİK RAPOR  —  FABRİKA BİRİMLERİ",
                            C["kirmizi"], satir=satir)
        satir += 1

        fabrika_kartlari = [
            ("⚗  Polisan Kimya Departmanı",  ozet["kimya"],  C["kimya"],
             lambda: self._goster_dinamik("Kimya"), None,
             lambda: self._dept_excel("Kimya")),
            ("🧪  Dow Departmanı",   ozet["dow"],    C["dow"],
             lambda: self._goster_dinamik("Dow"), None,
             lambda: self._dept_excel("Dow")),
            ("🔩  Diğer / Genel",    ozet["diger"],  "#5D4037",
             lambda: self._goster_dinamik("Diğer"), None,
             lambda: self._dept_excel("Diğer")),
        ]
        bos = self.motor.df_dinamik.empty
        for sutun, (baslik, (t,o,r), renk, cmd, altlar, exc) in enumerate(fabrika_kartlari):
            self._buyuk_kart(baslik=baslik, toplam=t, onay=o, red=r,
                             renk=renk, satir=satir, sutun=sutun,
                             veri_yok=bos, cmd=cmd, alt_birimler=altlar,
                             excel_cmd=exc)
        satir += 1

        # ── Terminal bölümü ──────────────────────────────────────────────────
        self._bolum_basligi("⚓  PREGATE SUNUM  —  TERMİNAL BİRİMLERİ",
                            C["mavi"], satir=satir)
        satir += 1

        terminal_altlar = [
            ("Pregate Red",   lambda: self._goster_alt_birim("Pregate Red")),
            ("Sevkiyat Red",  lambda: self._goster_alt_birim("Sevkiyat Red")),
            ("Operasyon Red", lambda: self._goster_alt_birim("Operasyon Red")),
        ]
        terminal_kartlari = [
            ("🏗  Poliport Terminal", ozet["terminal"], C["terminal"],
             lambda: self._goster_pregate("Poliport Terminal"),
             terminal_altlar, lambda: self._dept_excel("Poliport Terminal")),
            ("📦  Antrepo / Kuru Yük", ozet["antrepo"], C["antrepo"],
             lambda: self._goster_pregate("Antrepo"),
             None, lambda: self._dept_excel("Antrepo")),
        ]
        bos_pre = self.motor.df_pregate.empty
        # Poliport Terminal için Pregate/Sevkiyat/Operasyon Red kırılımı
        pol_red_detay = self.motor.pregate_red_detay_liste("Poliport Terminal") \
            if not bos_pre else None

        for sutun, (baslik, (t,o,r), renk, cmd, altlar, exc) in enumerate(terminal_kartlari):
            rd = pol_red_detay if baslik.startswith("🏗") else None
            self._buyuk_kart(baslik=baslik, toplam=t, onay=o, red=r,
                             renk=renk, satir=satir, sutun=sutun,
                             veri_yok=bos_pre, cmd=cmd, alt_birimler=altlar,
                             excel_cmd=exc, red_detay=rd)
        satir += 1

        # ── Kalite uyarıları ─────────────────────────────────────────────────
        self._kalite_uyarilari_goster(satir)

    def _gorsel_band_goster(self):
        """Ana sayfada tesis görsellerini yan yana gösterir."""
        if not (self._img_poliport or self._img_kuru):
            return

        # Başlık
        satir_no = 10  # Diğer kartların altına
        band_baslik = ctk.CTkLabel(
            self.dash,
            text="TESİS GÖRSELLERİ  —  Poliport Terminal",
            font=("Arial", 12, "bold"),
            text_color=C["metin_koyu"],
            fg_color="transparent",
            anchor="w"
        )
        band_baslik.grid(row=satir_no, column=0, columnspan=3,
                         sticky="w", padx=14, pady=(22, 6))

        gorsel_cerceve = ctk.CTkFrame(self.dash, fg_color="transparent")
        gorsel_cerceve.grid(row=satir_no+1, column=0, columnspan=3,
                            sticky="ew", padx=10, pady=(0, 8))
        gorsel_cerceve.grid_columnconfigure((0, 1), weight=1)

        gorsel_cfg = [
            (self._img_poliport, "Poliport Sıvı Terminali",   0),
            (self._img_kuru,     "Kuru Yük Terminali",        1),
        ]
        for img, baslik, sutun in gorsel_cfg:
            if not img:
                continue
            kart = ctk.CTkFrame(
                gorsel_cerceve,
                fg_color=C["card"],
                corner_radius=10,
                border_width=1, border_color=C["border"]
            )
            kart.grid(row=0, column=sutun, padx=8, sticky="nsew")
            ctk.CTkLabel(kart, image=img, text="",
                         corner_radius=8
                         ).pack(padx=8, pady=(8, 4))
            ctk.CTkLabel(kart, text=baslik,
                         font=("Arial", 11), text_color=C["metin_dim"]
                         ).pack(pady=(0, 8))

        # İmza bandı — her sayfanın en altında
        imza_band = ctk.CTkFrame(
            self.dash,
            fg_color=C["card2"],
            corner_radius=8
        )
        imza_band.grid(row=satir_no+2, column=0, columnspan=3,
                       sticky="ew", padx=10, pady=(6, 10))
        imza_band.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            imza_band,
            text=f"Bu program  {HAZIRLAYEN}  tarafından tasarlanmış ve geliştirilmiştir.   "
                 f"|   Seymen Raporlama  {UYGULAMA_SURUMU}   |   Pregate Operasyon Ekibi",
            font=("Arial", 10, "italic"),
            text_color=C["metin_koyu"],
            anchor="center"
        ).grid(row=0, column=0, pady=10, padx=20)

    def _imza_footer_ekle(self, parent):
        """Her rapor sayfasının sonuna imza ekler."""
        self._yaz("\n", None)
        self._yaz(
            f"  ─────────────────────────────────────────────────────────────\n",
            "uyari")
        self._yaz(
            f"  Bu rapor  {HAZIRLAYEN}  tarafından tasarlanmış Seymen Raporlama "
            f"sistemi ile oluşturulmuştur.   {UYGULAMA_SURUMU}\n",
            "basarili")

    def _bolum_basligi(self, metin, renk, satir):
        ctk.CTkLabel(
            self.dash, text=metin,
            font=("Arial", 11, "bold"),
            text_color=renk,
            fg_color=C["card2"],
            corner_radius=6
        ).grid(row=satir, column=0, columnspan=3,
               sticky="ew", padx=8, pady=(6, 2))

    def _buyuk_kart(self, baslik, toplam, onay, red,
                    renk, satir, sutun,
                    veri_yok=False, cmd=None, alt_birimler=None,
                    excel_cmd=None, red_detay=None):
        kart = ctk.CTkFrame(
            self.dash,
            fg_color=C["card"],
            corner_radius=8,
            border_width=1,
            border_color=renk if not veri_yok else C["border"]
        )
        kart.grid(row=satir, column=sutun,
                  padx=5, pady=3, sticky="nsew")

        # Başlık şeridi — 28px
        baslik_serit = ctk.CTkFrame(kart, fg_color=renk if not veri_yok
                                    else C["border"],
                                    corner_radius=6, height=28)
        baslik_serit.pack(fill="x")
        baslik_serit.pack_propagate(False)
        baslik_serit.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(
            baslik_serit, text=baslik,
            font=("Arial", 10, "bold"),
            text_color="#FFFFFF" if not veri_yok else C["metin_dim"]
        ).grid(row=0, column=0, padx=8, pady=4, sticky="w")

        if excel_cmd and not veri_yok:
            ctk.CTkButton(
                baslik_serit, text="📥",
                width=20, height=18,
                fg_color="transparent", hover_color="#444444",
                text_color="#FFFFFF", font=("Arial", 11),
                command=excel_cmd
            ).grid(row=0, column=1, padx=(0,4), pady=4, sticky="e")

        if veri_yok:
            ctk.CTkLabel(kart, text="⚠  Veri yüklenmedi",
                         font=("Arial", 9), text_color=C["turuncu"]
                         ).pack(pady=6)
        else:
            # Sayılar — tek satır, kompakt
            say = ctk.CTkFrame(kart, fg_color="transparent")
            say.pack(fill="x", padx=8, pady=(4, 0))
            say.grid_columnconfigure((0,1,2), weight=1)
            for idx, (lbl, val, clr) in enumerate([
                ("Toplam", toplam, C["mavi"]),
                ("Onay",   onay,   C["yesil"]),
                ("Red",    red,    C["kirmizi"]),
            ]):
                ctk.CTkLabel(say, text=lbl, font=("Arial", 8),
                             text_color=C["metin_dim"]
                             ).grid(row=0, column=idx, sticky="ew")
                ctk.CTkLabel(say, text=str(val),
                             font=("Arial", 16, "bold"),
                             text_color=clr
                             ).grid(row=1, column=idx, sticky="ew")

            # Pregate/Sevkiyat/Operasyon kırılımı — Red kolonunun altında
            if red_detay:
                alt_f = ctk.CTkFrame(say, fg_color="transparent")
                alt_f.grid(row=2, column=2, sticky="ew", pady=(1,0))
                for tip, key, rk in [
                    ("Preg.", "Pregate Red",   "#E74C3C"),
                    ("Sevk.", "Sevkiyat Red",  "#E67E22"),
                    ("Oper.", "Operasyon Red", "#2E86DE"),
                ]:
                    ctk.CTkLabel(alt_f, text=f"{tip}: {red_detay.get(key,0)}",
                                 font=("Arial", 7, "bold"), text_color=rk
                                 ).pack(anchor="center")

            # Progress bar — ince
            if toplam > 0:
                oran = onay / toplam
                pb = ctk.CTkProgressBar(kart, progress_color=C["yesil"],
                                        fg_color=C["border"],
                                        height=5, corner_radius=2)
                pb.pack(fill="x", padx=8, pady=(3, 0))
                pb.set(oran)
                ctk.CTkLabel(kart,
                             text=f"Onay Oranı: %{round(oran*100,1)}",
                             font=("Arial", 8), text_color=C["metin_dim"]
                             ).pack(pady=(1, 2))

        # Detay butonu
        if cmd:
            ctk.CTkButton(kart, text="Detayları Gör  ›",
                          fg_color=renk if not veri_yok else C["border"],
                          hover_color="#444",
                          font=("Arial", 10), height=22,
                          command=cmd
                          ).pack(fill="x", padx=8, pady=(2, 2))

        # Alt birim butonları (Pregate/Sevkiyat/Operasyon)
        if alt_birimler and not veri_yok:
            ctk.CTkFrame(kart, height=1, fg_color=C["border"]
                         ).pack(fill="x", padx=8)
            for ab_txt, ab_cmd in alt_birimler:
                ctk.CTkButton(kart, text=f"  └  {ab_txt}",
                              fg_color="#0B3028", hover_color="#0F4038",
                              font=("Arial", 9), height=20, anchor="w",
                              command=ab_cmd
                              ).pack(fill="x", padx=8, pady=1)

        ctk.CTkFrame(kart, height=1, fg_color=C["border"]
                     ).pack(fill="x", padx=8, pady=2)

    def _kalite_uyarilari_goster(self, baslangic_satir=4):
        uyarilar = []
        if not self.motor.df_dinamik.empty:
            eks = self.motor.eksik_analiz(
                self.motor.df_dinamik, ["nakliye","havuz","emniyet"])
            for col, sayi in eks.items():
                uyarilar.append(f"  🏭  Dinamik  /  '{col}'  →  {sayi} boş satır")
        if not self.motor.df_pregate.empty:
            eks2 = self.motor.eksik_analiz(
                self.motor.df_pregate, ["nakliye","durum","aciklama"])
            for col, sayi in eks2.items():
                uyarilar.append(f"  ⚓  Pregate  /  '{col}'  →  {sayi} boş satır")

        if not uyarilar:
            return

        self._bolum_basligi("⚠  VERİ KALİTE UYARILARI",
                            C["turuncu"], satir=baslangic_satir)
        uyari_kart = ctk.CTkFrame(self.dash, fg_color=C["card"],
                                   corner_radius=10,
                                   border_width=1, border_color=C["turuncu"])
        uyari_kart.grid(row=baslangic_satir+1, column=0, columnspan=3,
                        sticky="ew", padx=10, pady=4)
        for u in uyarilar:
            ctk.CTkLabel(uyari_kart, text=u, font=("Arial", 10),
                         text_color=C["turuncu"], anchor="w"
                         ).pack(anchor="w", padx=14, pady=3)

    # =========================================================================
    # DİNAMİK RAPOR SAYFASI
    # =========================================================================
    def _goster_dinamik(self, dep, alt_birim=None):
        if self.motor.df_dinamik.empty:
            messagebox.showwarning("Uyarı", "Dinamik Rapor dosyası yüklü değil.\n"
                                   "Sol menüden 'Dinamik Rapor Yükle' butonunu kullanın.")
            return

        DEP_ADLARI = {
            "Kimya":  "Polisan Kimya Departmanı",
            "Dow":    "Dow Departmanı",
            "Diğer":  "Diğer / Genel",
        }
        dep_adi = DEP_ADLARI.get(dep, dep)

        if alt_birim:
            baslik = f"🏭  {dep_adi}  ›  {alt_birim}"
        else:
            baslik = f"🏭  Dinamik Rapor  —  {dep_adi}"
        self._ekran_hazirla(baslik)

        df = self.motor._dinamik_filtrele(dep, alt_birim)
        t, o, r = self.motor.dinamik_kpi(df)
        self._kpi_guncelle(t, o, r)

        self._yaz(f"  {baslik.upper().replace('🏭  ','')}  —  ÖZET  \n", "h1")
        self._yaz(self._fmt("Toplam Kayıt",  t, s="▶"), "tek")
        self._yaz(self._fmt("Onaylanan",     o, s="▶"), "cift")
        self._yaz(self._fmt("Reddedilen",    r, s="▶"), "tek")
        self._yaz("\n")

        # Alt birimler (sadece ana departman görünümünde göster)
        if not alt_birim:
            alt = self.motor.dinamik_havuz_detay(dep)
            if alt:
                self._yaz("  ALT BİRİM  (HAVUZ)  DETAYLARI  \n", "h1")
                for i, (pool, pt, po, pr) in enumerate(alt):
                    tag = "tek" if i % 2 == 0 else "cift"
                    self._yaz(self._fmt(pool, pt, po, pr), tag)
                self._yaz("\n")

        # Top nakliyeci
        self._yaz("  EN ÇOK ARAÇ GÖNDEREN NAKLİYECİLER  (İLK 10)  \n", "h1")
        nak_col = sutun_bul(df, ["nakliyeci adi","nakliyeci adı","nakliye"])
        if nak_col and not df.empty:
            top_nak = df[nak_col].astype(str).str.strip().value_counts().head(10)
            for i, (ad, s) in enumerate(top_nak.items()):
                tag = "tek" if i % 2 == 0 else "cift"
                self._yaz(self._fmt(ad, f"{s} Araç", s="✔"), tag)
        else:
            self._yaz("  ⚠  Nakliyeci sütunu bulunamadı veya veri yok.\n", "uyari")
        self._yaz("\n")

        # Top nedenler
        self._yaz("  BAŞLICA RED / İPTAL NEDENLERİ  (İLK 10)  \n", "h1")
        neden_col = sutun_bul(df, ["iptal neden","neden","aciklama"])
        if neden_col and not df.empty:
            top_neden = df[neden_col].astype(str).str.strip().value_counts().head(10)
            for i, (n, s) in enumerate(top_neden.items()):
                tag = "tek" if i % 2 == 0 else "cift"
                self._yaz(self._fmt(n, f"{s} Kez", s="✖"), tag)
        else:
            self._yaz("  ⚠  Red nedeni sütunu bulunamadı.\n", "uyari")

        self._eksik_uyari_ekle(df)
        self._imza_footer_ekle(None)

    # =========================================================================
    # PREGATE SAYFASI
    # =========================================================================
    def _goster_pregate(self, dep):
        if self.motor.df_pregate.empty:
            messagebox.showwarning("Uyarı", "Pregate Sunum dosyası yüklü değil.")
            return
        self._ekran_hazirla(f"⚓  Pregate Sunum  —  {dep}", mod="ozel")

        df = self.motor._pregate_filtrele(dep)
        t, o, r = self.motor.pregate_kpi(df)
        # Poliport Terminal için Pregate/Sevkiyat/Operasyon Red ayrımı
        if dep == "Poliport Terminal":
            red_detay = self.motor.pregate_red_detay_liste(dep)
            self._kpi_guncelle(t, o, r, red_detay=red_detay)
        else:
            self._kpi_guncelle(t, o, r)

        ana = self._ozel_frame
        ana.grid_rowconfigure(1, weight=1)
        ana.grid_columnconfigure(0, weight=1)

        # ── Üst buton satırı ─────────────────────────────────────────────────
        btn_f = ctk.CTkFrame(ana, fg_color="transparent")
        btn_f.grid(row=0, column=0, sticky="ew", pady=(0, 6))

        if dep == "Poliport Terminal":
            ctk.CTkButton(
                btn_f,
                text="🔍  Tüm Red Kayıtları — Detay Popup (Filtreli)",
                fg_color=C["kirmizi"], hover_color="#7A1A2A",
                font=("Arial", 12, "bold"), height=36, corner_radius=8,
                command=lambda: self._pregate_tum_red_popup(df)
            ).pack(side="left", padx=(0, 8))

            ctk.CTkButton(
                btn_f,
                text="📊  Nakliyeci Red Matrix",
                fg_color="#5D4037", hover_color="#3D2410",
                font=("Arial", 12, "bold"), height=36, corner_radius=8,
                command=self._nakliye_red
            ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_f,
            text="📥  Excel İndir",
            fg_color="#7D6608", hover_color="#5D4E08",
            text_color=C["altin"],
            font=("Arial", 12, "bold"), height=36, corner_radius=8,
            command=lambda: self._firma_excele_kaydet(df, dep)
        ).pack(side="left")

        # ── Metin özet alanı ─────────────────────────────────────────────────
        txt = ctk.CTkTextbox(
            ana, font=("Consolas", 12),
            fg_color=C["card"], text_color=C["metin"],
            corner_radius=10, border_width=1, border_color=C["border"]
        )
        txt.grid(row=1, column=0, sticky="nsew")
        t2 = txt._textbox
        h1_bg = C["tag_h1_bg"]; tek_bg = C["tag_tek_bg"]; cift_bg = C["tag_cift_bg"]
        t2.tag_configure("h1",   background=h1_bg,   foreground=C["altin"],
                         font=("Arial",14,"bold"), spacing1=12, spacing3=12)
        t2.tag_configure("h2",   foreground=C["mavi"],
                         font=("Arial",12,"bold"), spacing1=8, spacing3=4)
        t2.tag_configure("tek",  background=tek_bg,  foreground=C["metin"],
                         font=("Consolas",12), spacing1=5, spacing3=5, tabs=("460","620","780"))
        t2.tag_configure("cift", background=cift_bg, foreground=C["metin"],
                         font=("Consolas",12), spacing1=5, spacing3=5, tabs=("460","620","780"))
        t2.tag_configure("uyari", foreground=C["turuncu"],
                         font=("Arial",12,"bold"), spacing1=6)

        def yaz(metin, stil=None):
            if stil: t2.insert("end", metin, stil)
            else: txt.insert("end", metin)

        kategoriler, top_nak, top_acik = self.motor.pregate_red_detay(dep)

        yaz(f"  {dep.upper()}  —  BİRİM BAZLI RED ANALİZİ  \n", "h1")
        for i, (kat, sayi) in enumerate(kategoriler.items()):
            yaz(f" ▶ {kat.upper()[:50]}\t: {sayi}\n", "tek" if i%2==0 else "cift")
        yaz("\n")

        yaz("  GENEL İSTATİSTİK  \n", "h1")
        yaz(f" ▶ TOPLAM ARAÇ\t: {t}\n", "tek")
        yaz(f" ▶ ONAYLANAN\t: {o}\n", "cift")
        yaz(f" ▶ TOPLAM RED\t: {r}\n", "tek")
        yaz("\n")

        yaz("  EN ÇOK RED EDİLEN NAKLİYECİLER  (İLK 10)  \n", "h1")
        if top_nak:
            for i, (ad, s) in enumerate(top_nak):
                yaz(f" ✖ {str(ad).upper()[:50]}\t: {s} Red\n",
                    "tek" if i%2==0 else "cift")
        else:
            yaz("  ⚠  Nakliyeci sütunu bulunamadı.\n", "uyari")
        yaz("\n")

        yaz("  BAŞLICA İPTAL AÇIKLAMALARI  (İLK 10)  \n", "h1")
        if top_acik:
            for i, (n, s) in enumerate(top_acik):
                yaz(f" ▪ {str(n)[:50]}\t: {s} Kez\n",
                    "tek" if i%2==0 else "cift")
        else:
            yaz("  ⚠  Açıklama sütunu bulunamadı.\n", "uyari")

    def _pregate_tum_red_popup(self, df_tum):
        """Poliport Terminal tüm red kayıtları — nakliyeci+plaka filtreli popup."""
        if df_tum is None or df_tum.empty:
            messagebox.showwarning("Uyarı","Veri yok."); return

        adim_col_r = sutun_bul(df_tum, ["iptal adimi", "iptal adımı", "İptal Adımı"])
        if not adim_col_r:
            messagebox.showwarning("Uyarı","İptal Adımı sütunu bulunamadı."); return

        sa_r = df_tum[adim_col_r].astype(str).str.strip()
        # İptal Adımı'ndan red kayıtları
        df_red = df_tum[sa_r.isin(TUM_RED_ADIMLARI)].copy()

        if df_red.empty:
            messagebox.showinfo("Bilgi","Red kaydı bulunamadı."); return

        pop = ctk.CTkToplevel(self)
        pop.title(f"Poliport Terminal  —  Tüm Red Kayıtları  ({len(df_red)} kayıt)")
        pop.geometry("1200x720")
        pop.resizable(True, True)
        pop.configure(fg_color=C["bg"])
        pop.grab_set(); pop.lift(); pop.focus_force()

        # Sütunları normalize et
        df = df_red.copy()
        df.columns = [str(c).replace("\n"," ").strip() for c in df.columns]
        tum = [c for c in df.columns if c not in ("nan","NaT","None","")]

        nak_col   = sutun_bul(df, ["nakliyeci adi","nakliyeci adı","nakliye"])
        plaka_col = sutun_bul(df, ["çekici","cekici","plaka"])

        # Nakliyeci listesi
        nak_listesi = ["— Tümü —"]
        if nak_col:
            nak_listesi += sorted(df[nak_col].dropna().astype(str).str.strip().unique().tolist())

        # Varsayılan sütunlar
        oncelik = ["Kayıt Tarihi","Çekici","Dorse Plaka","Tank Kodu","Nakliyeci",
                   "Durum","İptal Aciklama","Kantar Red","İptal Eden","Departman"]
        varsayilan = []
        for o in oncelik:
            eslesme = next((c for c in tum if o.lower().replace(" ","") in c.lower().replace(" ","")), None)
            if eslesme and eslesme not in varsayilan:
                varsayilan.append(eslesme)
        if not varsayilan:
            varsayilan = tum[:6]

        secim = {c: ctk.BooleanVar(value=(c in varsayilan)) for c in tum}
        nak_var   = ctk.StringVar(value="— Tümü —")
        plaka_var = ctk.StringVar(value="")
        filtre_var = ctk.StringVar(value="")
        tablo_ref = [None]

        # Ana grid
        pop.grid_rowconfigure(1, weight=1)
        pop.grid_columnconfigure(1, weight=1)

        # Başlık
        bb = ctk.CTkFrame(pop, fg_color="#7B1010", corner_radius=0, height=44)
        bb.grid(row=0, column=0, columnspan=2, sticky="ew")
        bb.pack_propagate(False)
        ctk.CTkLabel(bb, text=f"  🔍  Poliport Terminal  —  {len(df_red)} Red Kayıt",
                     font=("Arial",13,"bold"), text_color="#FFFFFF").pack(side="left", padx=10)
        ctk.CTkButton(bb, text="✕  Kapat", fg_color="#500A0A", hover_color="#3A0606",
                      text_color="#FFFFFF", font=("Arial",11), width=90, height=30,
                      command=pop.destroy).pack(side="right", padx=6, pady=7)
        ctk.CTkButton(bb, text="📥  Excel", fg_color="#7D6608", hover_color="#5D4E08",
                      text_color=C["altin"], font=("Arial",11), width=80, height=30,
                      command=lambda: self._firma_excele_kaydet(df_red, "Terminal_Red_Detay")
                      ).pack(side="right", padx=4, pady=7)
        ctk.CTkButton(bb, text="📋  Kopyala", fg_color="#1A3A5C", hover_color="#0F2540",
                      text_color="#FFFFFF", font=("Arial",11), width=90, height=30,
                      command=lambda: (self._genel_tablo_kopyala(tablo_ref[0], True)
                                       if tablo_ref[0] else None)
                      ).pack(side="right", padx=4, pady=7)

        # Sol: sütun seçici
        sol = ctk.CTkFrame(pop, fg_color=C["card"], corner_radius=0, width=210)
        sol.grid(row=1, column=0, sticky="nsew"); sol.grid_propagate(False)
        sol.grid_rowconfigure(2, weight=1); sol.grid_columnconfigure(0, weight=1)
        ctk.CTkLabel(sol, text="  Sütunlar", font=("Arial",11,"bold"),
                     text_color=C["altin"], anchor="w"
                     ).grid(row=0, column=0, sticky="ew", padx=8, pady=(8,4))
        hepsi = ctk.CTkFrame(sol, fg_color="transparent")
        hepsi.grid(row=1, column=0, sticky="ew", padx=6, pady=(0,4))
        hepsi.grid_columnconfigure((0,1), weight=1)
        ctk.CTkButton(hepsi, text="✅ Tümü", fg_color=C["yesil"], hover_color="#155230",
                      font=("Arial",9), height=24,
                      command=lambda: [v.set(True) for v in secim.values()]
                      ).grid(row=0, column=0, sticky="ew", padx=2)
        ctk.CTkButton(hepsi, text="☐ Kaldır", fg_color=C["card2"], hover_color=C["border"],
                      font=("Arial",9), height=24,
                      command=lambda: [v.set(False) for v in secim.values()]
                      ).grid(row=0, column=1, sticky="ew", padx=2)
        cb_frame = ctk.CTkScrollableFrame(sol, fg_color="transparent",
                                          scrollbar_button_color=C["border"])
        cb_frame.grid(row=2, column=0, sticky="nsew", padx=4, pady=4)
        for col in tum:
            ctk.CTkCheckBox(cb_frame, text=col[:24]+("…" if len(col)>24 else ""),
                            variable=secim[col], font=("Arial",10),
                            text_color=C["metin"], fg_color=C["mavi"],
                            hover_color="#1A5C9A", checkmark_color="#FFFFFF",
                            border_color=C["border"], height=24
                            ).pack(anchor="w", padx=6, pady=2)
        ctk.CTkButton(sol, text="✓  Uygula", fg_color=C["mavi"], hover_color="#1460A0",
                      font=("Arial",11,"bold"), height=34,
                      command=lambda: _guncelle()
                      ).grid(row=3, column=0, sticky="ew", padx=6, pady=6)

        # Sağ: filtre + tablo
        sag = ctk.CTkFrame(pop, fg_color="transparent", corner_radius=0)
        sag.grid(row=1, column=1, sticky="nsew")
        sag.grid_rowconfigure(1, weight=1); sag.grid_columnconfigure(0, weight=1)

        filtre_f = ctk.CTkFrame(sag, fg_color=C["card2"], corner_radius=0, height=50)
        filtre_f.grid(row=0, column=0, sticky="ew"); filtre_f.pack_propagate(False)

        ctk.CTkLabel(filtre_f, text="  Nakliyeci:", font=("Arial",10),
                     text_color=C["metin_dim"]).pack(side="left", padx=(8,2))
        nak_om = ctk.CTkOptionMenu(filtre_f, variable=nak_var,
                                   values=nak_listesi[:100], width=200, height=28,
                                   font=("Arial",9), fg_color=C["card"],
                                   button_color=C["border"], dropdown_fg_color=C["card"],
                                   command=lambda v: _guncelle())
        nak_om.pack(side="left", padx=4, pady=10)

        ctk.CTkLabel(filtre_f, text="Plaka/Metin:", font=("Arial",10),
                     text_color=C["metin_dim"]).pack(side="left", padx=(4,2))
        ara_e = ctk.CTkEntry(filtre_f, textvariable=filtre_var,
                             placeholder_text="Plaka, açıklama...",
                             width=180, height=28, font=("Arial",10))
        ara_e.pack(side="left", padx=4)
        ara_e.bind("<KeyRelease>", lambda e: _guncelle())

        ctk.CTkButton(filtre_f, text="🔄", fg_color=C["card2"], hover_color=C["border"],
                      text_color=C["metin_dim"], font=("Arial",10), width=30, height=28,
                      command=lambda: (nak_var.set("— Tümü —"), filtre_var.set(""), _guncelle())
                      ).pack(side="left", padx=4)

        lbl_sayac = ctk.CTkLabel(filtre_f, text="", font=("Arial",9),
                                  text_color=C["metin_koyu"])
        lbl_sayac.pack(side="left", padx=8)

        tablo_c = ctk.CTkFrame(sag, fg_color=C["card"], corner_radius=0)
        tablo_c.grid(row=1, column=0, sticky="nsew")
        tablo_c.grid_rowconfigure(0, weight=1); tablo_c.grid_columnconfigure(0, weight=1)
        sy = ttk.Scrollbar(tablo_c, orient="vertical"); sy.grid(row=0, column=1, sticky="ns")
        sx = ttk.Scrollbar(tablo_c, orient="horizontal"); sx.grid(row=1, column=0, sticky="ew")
        tablo = ttk.Treeview(tablo_c, yscrollcommand=sy.set, xscrollcommand=sx.set, show="headings")
        tablo.grid(row=0, column=0, sticky="nsew")
        sy.configure(command=tablo.yview); sx.configure(command=tablo.xview)
        tablo_ref[0] = tablo

        from tkinter import Menu as TkM
        m = TkM(pop, tearoff=0, bg="#1A2438", fg="white",
                activebackground=C["mavi"], activeforeground="white", font=("Arial",10))
        m.add_command(label="📋  Seçili Kopyala", command=lambda: self._genel_tablo_kopyala(tablo, False))
        m.add_command(label="📄  Tümünü Kopyala", command=lambda: self._genel_tablo_kopyala(tablo, True))
        m.add_command(label="📥  Excel İndir", command=lambda: self._firma_excele_kaydet(df_red, "Terminal_Red"))
        def _stk(e):
            item = tablo.identify_row(e.y)
            if item: tablo.selection_set(item)
            try: m.tk_popup(e.x_root, e.y_root)
            finally: m.grab_release()
        tablo.bind("<Button-3>", _stk); tablo.bind("<Button-2>", _stk)

        def _guncelle():
            df_f = df.copy()
            # Nakliyeci filtresi
            if nak_var.get() != "— Tümü —" and nak_col and nak_col in df_f.columns:
                df_f = df_f[df_f[nak_col].astype(str).str.strip() == nak_var.get()]
            # Metin filtresi
            ara = filtre_var.get().strip().lower()
            if ara:
                arama_kols = [c for c in [plaka_col, nak_col,
                              sutun_bul(df_f, ["aciklama","iptal ac","kantar"])]
                              if c and c in df_f.columns]
                if arama_kols:
                    mask = df_f.apply(lambda row: any(
                        ara in str(row[c]).lower() for c in arama_kols), axis=1)
                    df_f = df_f[mask]

            secili = [c for c in tum if secim[c].get()]
            if not secili: secili = varsayilan

            tablo.delete(*tablo.get_children())
            tablo["columns"] = secili
            gen = {"kayıttarihi":120,"çekici":85,"dorseplaka":85,"tankodu":80,
                   "nakliyeci":180,"durum":110,"aciklama":200,"kantarred":180,
                   "iptaeden":120,"departman":120}
            for c in secili:
                ck = c.lower().replace(" ","")
                w = next((v for k,v in gen.items() if k in ck), 100)
                tablo.column(c, anchor="w", width=w, minwidth=50)
                tablo.heading(c, text=c, anchor="w")

            for i, (_, row) in enumerate(df_f.head(500).iterrows()):
                vals = []
                for c in secili:
                    val = row.get(c,"")
                    try:
                        if pd.notna(val) and hasattr(val,"strftime"):
                            val = val.strftime("%d.%m.%Y %H:%M")
                    except Exception: pass
                    s = str(val)
                    vals.append(s[:100] if s not in ("nan","NaT","None") else "")
                tablo.insert("","end",values=vals,tags=("tek_r" if i%2==0 else "cift_r",))

            if TEMA_MOD == "Light":
                tablo.tag_configure("tek_r",  background="#F7FAFC", foreground="#1A202C", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#EDF2F7", foreground="#1A202C", font=("Consolas",10))
            else:
                tablo.tag_configure("tek_r",  background="#141C2B", foreground="#E8EDF5", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#111827", foreground="#E8EDF5", font=("Consolas",10))

            lbl_sayac.configure(text=f"  {len(df_f)}/{len(df)} kayıt  |  {len(secili)} sütun")

        _guncelle()

    # =========================================================================
    # PREGATE ALT BİRİM (Red Tipi Detay)
    # =========================================================================
    def _goster_alt_birim(self, red_tipi):
        if self.motor.df_pregate.empty:
            messagebox.showwarning("Uyarı", "Pregate dosyası yüklü değil.")
            return

        self._ekran_hazirla(
            f"⚓  Poliport Terminal  ›  {red_tipi}  Detay Analizi",
            mod="ozel")

        df_tum = self.motor._pregate_filtrele("Poliport Terminal")
        t, o, r = self.motor.pregate_kpi(df_tum)
        self._kpi_guncelle(t, o, r)

        # Sütunları bul
        durum_col = sutun_bul(df_tum, ["durum"])
        nak_col   = sutun_bul(df_tum, ["nakliyeci adi","nakliyeci adı","nakliye"])
        tarih_col = sutun_bul(df_tum, ["tarih","date","giriş","giris"])
        plaka_col = sutun_bul(df_tum, ["plaka","plate","çekici","cekici"])
        acik_col  = sutun_bul(df_tum, ["iptal aciklama","aciklama","neden","kantar red","kantar"])

        # Red DataFrame — direkt Durum'a göre, personel ayrımı yok
        df_red = self.motor.pregate_red_df("Poliport Terminal", red_tipi)
        red_say = len(df_red)

        # Sol: Nakliyeci listesi
        firma_listesi = []
        if nak_col and not df_red.empty:
            s = df_red[nak_col].astype(str).str.strip()
            firma_listesi = list(s[s != "nan"].value_counts().head(20).items())

        # Sağ: İptal Açıklaması listesi (red tipine göre doğru sütun)
        acik_listesi = []
        if not df_red.empty:
            # Operasyon Red → Kantar Red sütunu, diğerleri → Ana Neden / İptal Aciklama
            if red_tipi == "Operasyon Red":
                ac = sutun_bul(df_red, ["kantar red","kantar"])
            else:
                ac = sutun_bul(df_red, ["ana neden","Ana Neden","iptal aciklama","aciklama","neden"])
            if ac:
                s = df_red[ac].astype(str).str.strip()
                acik_listesi = list(s[s != "nan"].value_counts().head(20).items())

        # Layout
        ana = self._ozel_frame
        ana.grid_rowconfigure(1, weight=1)
        ana.grid_columnconfigure(0, weight=1)

        # Üst buton satırı
        btn_f = ctk.CTkFrame(ana, fg_color="transparent")
        btn_f.grid(row=0, column=0, sticky="ew", pady=(0, 6))

        ctk.CTkButton(
            btn_f,
            text="🔍  Detay Tablosu  —  Plaka / Nakliyeci / İptal Nedenleri",
            fg_color=C["mavi"], hover_color="#1460A0",
            font=("Arial", 12, "bold"), height=34, corner_radius=8,
            command=lambda: self._red_detay_popup(
                red_tipi, df_red, nak_col, durum_col,
                acik_col, tarih_col, plaka_col)
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_f,
            text="📥  Excel İndir",
            fg_color="#7D6608", hover_color="#5D4E08",
            text_color=C["altin"],
            font=("Arial", 12, "bold"), height=34, corner_radius=8,
            command=lambda: self._firma_excele_kaydet(df_red, red_tipi)
        ).pack(side="left", padx=(0, 8))

        ctk.CTkButton(
            btn_f,
            text="📊  Pivot / Matrix",
            fg_color="#5D4037", hover_color="#3D2410",
            font=("Arial", 12, "bold"), height=34, corner_radius=8,
            command=self._nakliye_red
        ).pack(side="left")

        # İki panel: Nakliyeci | İptal Açıklaması
        panel = ctk.CTkFrame(ana, fg_color="transparent")
        panel.grid(row=1, column=0, sticky="nsew")
        panel.grid_rowconfigure(0, weight=1)
        panel.grid_columnconfigure(0, weight=1)
        panel.grid_columnconfigure(1, weight=1)

        baslik_sol = f"🏢  Nakliyeci Firma  ({red_say} {red_tipi})"
        baslik_sag = "📋  İptal Açıklaması  (Sebep Bazlı Adet)"

        def _liste_p(parent, sutun, baslik_txt, liste, val_clr, sembol):
            c = ctk.CTkFrame(parent, fg_color=C["card"], corner_radius=10,
                             border_width=1, border_color=C["border"])
            c.grid(row=0, column=sutun, sticky="nsew",
                   padx=(0,6) if sutun==0 else (6,0))
            c.grid_columnconfigure(0, weight=1)
            c.grid_rowconfigure(1, weight=1)
            ctk.CTkLabel(c, text=baslik_txt,
                         font=("Arial", 11, "bold"),
                         text_color=C["altin"], anchor="w"
                         ).grid(row=0, column=0, sticky="w", padx=10, pady=(8,4))
            sc = ctk.CTkScrollableFrame(c, fg_color="transparent",
                                        scrollbar_button_color=C["border"])
            sc.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0,8))
            sc.grid_columnconfigure(0, weight=1)
            if liste:
                for i, (ad, s) in enumerate(liste):
                    bg = C["card2"] if i%2==0 else "transparent"
                    rw = ctk.CTkFrame(sc, fg_color=bg, corner_radius=5)
                    rw.pack(fill="x", pady=1)
                    rw.grid_columnconfigure(0, weight=1)
                    ctk.CTkLabel(rw, text=f"  {sembol}  {str(ad)[:45]}",
                                 font=("Consolas", 10),
                                 text_color=C["metin"], anchor="w"
                                 ).grid(row=0, column=0, sticky="w", padx=4, pady=3)
                    ctk.CTkLabel(rw, text=f"{s}",
                                 font=("Arial", 10, "bold"),
                                 text_color=val_clr
                                 ).grid(row=0, column=1, padx=8)
            else:
                ctk.CTkLabel(sc, text="Bu tipte kayıt yok",
                             font=("Arial", 10),
                             text_color=C["metin_dim"]).pack(pady=14)

        _liste_p(panel, 0, baslik_sol,  firma_listesi, C["kirmizi"], "✖")
        _liste_p(panel, 1, baslik_sag, acik_listesi,  C["turuncu"], "▶")

    def _red_detay_popup(self, red_tipi, df_red,
                         nak_col, durum_col, acik_col, tarih_col, plaka_col):
        """Red tipinin tüm kayıtlarını gösteren detay popup — sütun seçici + filtre."""
        if df_red is None or df_red.empty:
            messagebox.showwarning("Uyarı", "Bu red tipinde kayıt bulunamadı.")
            return

        popup = ctk.CTkToplevel(self)
        popup.title(f"{red_tipi}  —  {len(df_red)} Kayıt")
        popup.geometry("1100x700")
        popup.resizable(True, True)
        popup.configure(fg_color=C["bg"])
        popup.grab_set(); popup.lift(); popup.focus_force()

        # DataFrame sütunlarını normalize et
        df = df_red.copy()
        df.columns = [str(c).replace("\n"," ").strip() for c in df.columns]
        tum_sutunlar = [c for c in df.columns if c not in ("nan","NaT","None","")]

        # Varsayılan sütunlar
        if red_tipi == "Operasyon Red":
            oncelik = ["Tarih","Çekici","Departman","İptal Adımı","Ana Neden",
                       "Kantar Red","Nakliyeci","İptal Eden"]
        else:
            oncelik = ["Tarih","Çekici","Departman","İptal Adımı","Ana Neden",
                       "Iptal Aciklama","Nakliyeci","İptal Eden"]
        varsayilan = []
        for o in oncelik:
            eslesme = next((c for c in tum_sutunlar if
                o.lower().replace(" ","") in c.lower().replace(" ","")), None)
            if eslesme and eslesme not in varsayilan:
                varsayilan.append(eslesme)
        if not varsayilan:
            varsayilan = tum_sutunlar[:6]

        secim = {c: ctk.BooleanVar(value=(c in varsayilan)) for c in tum_sutunlar}
        filtre_var = ctk.StringVar(value="")
        filtre_sutun_var = ctk.StringVar(value="— Tüm Sütunlar —")

        # Ana grid
        popup.grid_rowconfigure(1, weight=1)
        popup.grid_columnconfigure(1, weight=1)

        # Başlık
        bb = ctk.CTkFrame(popup, fg_color="#7B1010", corner_radius=0, height=44)
        bb.grid(row=0, column=0, columnspan=2, sticky="ew")
        bb.pack_propagate(False)
        ctk.CTkLabel(bb, text=f"  ✖  {red_tipi}  —  {len(df)} kayıt",
                     font=("Arial", 13, "bold"), text_color="#FFFFFF"
                     ).pack(side="left", padx=10)
        tablo_ref = [None]
        ctk.CTkButton(bb, text="✕  Kapat", fg_color="#500A0A", hover_color="#3A0606",
                      text_color="#FFFFFF", font=("Arial", 11), width=90, height=30,
                      command=popup.destroy).pack(side="right", padx=6, pady=7)
        ctk.CTkButton(bb, text="📥  Excel", fg_color="#7D6608", hover_color="#5D4E08",
                      text_color=C["altin"], font=("Arial", 11), width=80, height=30,
                      command=lambda: self._firma_excele_kaydet(df_red, red_tipi)
                      ).pack(side="right", padx=4, pady=7)
        ctk.CTkButton(bb, text="📋  Kopyala", fg_color="#1A3A5C", hover_color="#0F2540",
                      text_color="#FFFFFF", font=("Arial", 11), width=90, height=30,
                      command=lambda: (self._genel_tablo_kopyala(tablo_ref[0], True)
                                       if tablo_ref[0] else None)
                      ).pack(side="right", padx=4, pady=7)

        # Sol: sütun seçici
        sol = ctk.CTkFrame(popup, fg_color=C["card"], corner_radius=0, width=210)
        sol.grid(row=1, column=0, sticky="nsew")
        sol.grid_propagate(False)
        sol.grid_rowconfigure(2, weight=1)
        sol.grid_columnconfigure(0, weight=1)

        ctk.CTkLabel(sol, text="  Sütunlar", font=("Arial", 11, "bold"),
                     text_color=C["altin"], anchor="w"
                     ).grid(row=0, column=0, sticky="ew", padx=8, pady=(8,4))

        hepsi = ctk.CTkFrame(sol, fg_color="transparent")
        hepsi.grid(row=1, column=0, sticky="ew", padx=6, pady=(0,4))
        hepsi.grid_columnconfigure((0,1), weight=1)
        ctk.CTkButton(hepsi, text="✅ Tümü", fg_color=C["yesil"], hover_color="#155230",
                      font=("Arial", 9), height=24,
                      command=lambda: [v.set(True) for v in secim.values()]
                      ).grid(row=0, column=0, sticky="ew", padx=2)
        ctk.CTkButton(hepsi, text="☐ Kaldır", fg_color=C["card2"], hover_color=C["border"],
                      font=("Arial", 9), height=24,
                      command=lambda: [v.set(False) for v in secim.values()]
                      ).grid(row=0, column=1, sticky="ew", padx=2)

        cb_frame = ctk.CTkScrollableFrame(sol, fg_color="transparent",
                                          scrollbar_button_color=C["border"])
        cb_frame.grid(row=2, column=0, sticky="nsew", padx=4, pady=4)
        for col in tum_sutunlar:
            ctk.CTkCheckBox(cb_frame, text=col[:24]+("…" if len(col)>24 else ""),
                            variable=secim[col], font=("Arial", 10),
                            text_color=C["metin"], fg_color=C["mavi"],
                            hover_color="#1A5C9A", checkmark_color="#FFFFFF",
                            border_color=C["border"], height=24
                            ).pack(anchor="w", padx=6, pady=2)

        ctk.CTkButton(sol, text="✓  Uygula", fg_color=C["mavi"], hover_color="#1460A0",
                      font=("Arial", 11, "bold"), height=34,
                      command=lambda: _tabloyu_guncelle()
                      ).grid(row=3, column=0, sticky="ew", padx=6, pady=6)

        # Sağ: filtre + tablo
        sag = ctk.CTkFrame(popup, fg_color="transparent", corner_radius=0)
        sag.grid(row=1, column=1, sticky="nsew")
        sag.grid_rowconfigure(1, weight=1)
        sag.grid_columnconfigure(0, weight=1)

        filtre_band = ctk.CTkFrame(sag, fg_color=C["card2"], corner_radius=0, height=44)
        filtre_band.grid(row=0, column=0, sticky="ew")
        filtre_band.pack_propagate(False)

        ctk.CTkLabel(filtre_band, text="  Ara:", font=("Arial", 11),
                     text_color=C["metin_dim"]).pack(side="left", padx=(8,2))
        filtre_entry = ctk.CTkEntry(filtre_band, textvariable=filtre_var,
                                    width=200, height=30, font=("Arial", 11),
                                    fg_color=C["card"], border_color=C["border"])
        filtre_entry.pack(side="left", padx=4, pady=7)
        ctk.CTkLabel(filtre_band, text="  Sütunda:", font=("Arial", 11),
                     text_color=C["metin_dim"]).pack(side="left", padx=(4,2))
        sutun_menu = ctk.CTkOptionMenu(
            filtre_band, variable=filtre_sutun_var,
            values=["— Tüm Sütunlar —"] + tum_sutunlar,
            width=180, height=30, font=("Arial", 10),
            fg_color=C["card"], button_color=C["border"],
            dropdown_fg_color=C["card"],
            command=lambda _: _tabloyu_guncelle())
        sutun_menu.pack(side="left", padx=4)
        ctk.CTkButton(filtre_band, text="✕", fg_color=C["card2"], hover_color=C["border"],
                      text_color=C["metin_dim"], font=("Arial", 11), width=30, height=30,
                      command=lambda: (filtre_var.set(""), _tabloyu_guncelle())
                      ).pack(side="left", padx=2)
        filtre_entry.bind("<KeyRelease>", lambda e: _tabloyu_guncelle())

        tablo_cerceve = ctk.CTkFrame(sag, fg_color=C["card"], corner_radius=0)
        tablo_cerceve.grid(row=1, column=0, sticky="nsew")
        tablo_cerceve.grid_rowconfigure(0, weight=1)
        tablo_cerceve.grid_columnconfigure(0, weight=1)

        alt_band = ctk.CTkFrame(sag, fg_color=C["card2"], corner_radius=0, height=22)
        alt_band.grid(row=2, column=0, sticky="ew")
        alt_band.pack_propagate(False)
        lbl_sayac = ctk.CTkLabel(alt_band, text="", font=("Arial", 9),
                                  text_color=C["metin_koyu"], anchor="w")
        lbl_sayac.pack(side="left", padx=10)

        def _tabloyu_guncelle():
            secili = [c for c in tum_sutunlar if secim[c].get()]
            if not secili: secili = tum_sutunlar[:4]
            ara = filtre_var.get().strip().lower()
            ara_sutun = filtre_sutun_var.get()
            if ara:
                if ara_sutun == "— Tüm Sütunlar —":
                    mask = df.apply(lambda row: any(
                        ara in str(row[c]).lower() for c in secili if c in df.columns), axis=1)
                else:
                    mask = (df[ara_sutun].astype(str).str.lower().str.contains(ara, na=False)
                            if ara_sutun in df.columns else pd.Series([True]*len(df)))
                df_fil = df[mask]
            else:
                df_fil = df

            for w in tablo_cerceve.winfo_children(): w.destroy()
            sy2 = ttk.Scrollbar(tablo_cerceve, orient="vertical")
            sy2.grid(row=0, column=1, sticky="ns")
            sx2 = ttk.Scrollbar(tablo_cerceve, orient="horizontal")
            sx2.grid(row=1, column=0, sticky="ew")
            t = ttk.Treeview(tablo_cerceve, yscrollcommand=sy2.set,
                             xscrollcommand=sx2.set, show="headings")
            t.grid(row=0, column=0, sticky="nsew")
            sy2.configure(command=t.yview)
            sx2.configure(command=t.xview)
            tablo_ref[0] = t

            t["columns"] = secili
            gen = {"kayıttarihi":120,"plantarihi":110,"plaka":85,"çekici":85,
                   "nakliyeci":180,"durum":120,"aciklama":210,"kantarred":180,
                   "iptaledenkullanıcıadı":130,"iptaleden":120,"departman":130}
            for c in secili:
                ck = c.lower().replace(" ","")
                w = next((v for k,v in gen.items() if k in ck), 100)
                t.column(c, anchor="w", width=w, minwidth=50)
                t.heading(c, text=c.replace("\n"," "), anchor="w")

            for i, (_, row) in enumerate(df_fil.head(500).iterrows()):
                vals = []
                for c in secili:
                    val = row.get(c,"")
                    try:
                        if pd.notna(val) and hasattr(val,"strftime"):
                            val = val.strftime("%d.%m.%Y %H:%M")
                    except Exception: pass
                    s = str(val)
                    vals.append(s[:100] if s not in ("nan","NaT","None") else "")
                t.insert("","end",values=vals,tags=("tek_r" if i%2==0 else "cift_r",))

            if TEMA_MOD == "Light":
                t.tag_configure("tek_r",  background="#F7FAFC", foreground="#1A202C", font=("Consolas",10))
                t.tag_configure("cift_r", background="#EDF2F7", foreground="#1A202C", font=("Consolas",10))
            else:
                t.tag_configure("tek_r",  background="#141C2B", foreground="#E8EDF5", font=("Consolas",10))
                t.tag_configure("cift_r", background="#111827", foreground="#E8EDF5", font=("Consolas",10))

            from tkinter import Menu as TkM
            m = TkM(popup, tearoff=0, bg="#1A2438", fg="white",
                    activebackground=C["mavi"], activeforeground="white", font=("Arial",10))
            m.add_command(label="📋  Seçili Kopyala",
                          command=lambda: self._genel_tablo_kopyala(t, False))
            m.add_command(label="📄  Tümünü Kopyala",
                          command=lambda: self._genel_tablo_kopyala(t, True))
            m.add_command(label="📥  Excel İndir",
                          command=lambda: self._firma_excele_kaydet(df_red, red_tipi))
            def _stk(e, tbl=t, mn=m):
                item = tbl.identify_row(e.y)
                if item: tbl.selection_set(item)
                try: mn.tk_popup(e.x_root, e.y_root)
                finally: mn.grab_release()
            t.bind("<Button-3>", _stk); t.bind("<Button-2>", _stk)
            lbl_sayac.configure(
                text=f"  {len(df_fil)}/{len(df)} kayıt  |  {len(secili)} sütun  |  Sağ tık → menü")

        _tabloyu_guncelle()

    def _firma_excele_kaydet(self, df, ad):
        if df is None or df.empty:
            messagebox.showwarning("Uyarı","Veri yok."); return
        yol = filedialog.asksaveasfilename(
            title=f"{ad} Listesini Kaydet",
            defaultextension=".xlsx",
            initialfile=f"{ad.replace(' ','_')}_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not yol: return
        try:
            cols = [c for c in df.columns if not str(c).startswith("_")]
            df[cols].to_excel(yol, index=False)
            messagebox.showinfo("Başarılı", f"Kaydedildi:\n{yol}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")


    def _goster_ek(self):
        if self.motor.df_ek.empty:
            messagebox.showwarning("Uyarı", "Ek dosya yüklü değil.\n"
                                   "Sol menüden 'Ek Dosya Yükle' butonunu kullanın.")
            return
        self._ekran_hazirla("📄  Ek Dosya  —  Genel Analiz")
        df = self.motor.df_ek
        self._kpi_guncelle(len(df), 0, 0)

        self._yaz("  EK DOSYA  —  GENEL BİLGİ  \n", "h1")
        self._yaz(self._fmt("Toplam Satır",   len(df),         s="▶"), "tek")
        self._yaz(self._fmt("Sütun Sayısı",   len(df.columns), s="▶"), "cift")
        bos_toplam = int(df.isna().sum().sum())
        self._yaz(self._fmt("Toplam Boş Hücre", bos_toplam,    s="▶",
                             eksik=(bos_toplam > 0)),
                  "eksik" if bos_toplam > 0 else "tek")
        self._yaz("\n")

        self._yaz("  SÜTUN ANALİZİ  \n", "h1")
        for i, col in enumerate(df.columns):
            bos  = int(df[col].isna().sum())
            dolu = len(df) - bos
            tag  = "eksik" if bos > 0 else ("tek" if i % 2 == 0 else "cift")
            self._yaz(
                self._fmt(col, f"Dolu: {dolu}  /  Boş: {bos}",
                          eksik=(bos > 0)), tag)
        self._yaz("\n")

        self._yaz("  DEĞER DAĞILIMI  (Metin sütunları  —  İlk 5)  \n", "h1")
        for col in list(df.select_dtypes(include=["object"]).columns)[:8]:
            self._yaz(f"\n  {col.upper()}\n", "h2")
            top = df[col].astype(str).str.strip()
            top = top[top != "nan"].value_counts().head(5)
            for i, (v, s) in enumerate(top.items()):
                tag = "tek" if i % 2 == 0 else "cift"
                self._yaz(self._fmt(v, f"{s} kez", s="▪"), tag)

    # =========================================================================
    # NAKLİYECİ ANALİZİ
    # =========================================================================
    def _goster_nakliyeci(self):
        """Nakliyeci Analizi — Pregate verisi, nakliyeci dropdown + durum filtresi."""
        if self.motor.df_pregate.empty:
            messagebox.showwarning("Uyarı", "Önce Pregate Sunum dosyası yükleyin.")
            return

        self._ekran_hazirla("🚛  Nakliyeci Analizi  —  Gelişmiş Filtre", mod="ozel")

        # Sadece Pregate verisi — Terminal departmanı
        dep_col = sutun_bul(self.motor.df_pregate, ["departman"])
        if dep_col:
            sd_dep = self.motor.df_pregate[dep_col].astype(str).str.lower()
            self.motor.df_pol_tmp = self.motor.df_pregate[
                sd_dep.str.contains("terminal", na=False)].copy()
        else:
            self.motor.df_pol_tmp = self.motor.df_pregate.copy()
        df_tum = self.motor.df_pol_tmp.copy()

        # KPI — İptal Adımı sütunundan
        t, o, r = self.motor.pregate_kpi(df_tum)
        self._kpi_guncelle(t, o, r)

        # Nakliyeci listesi oluştur
        nak_col_g = sutun_bul(df_tum, ["nakliyeci adi","nakliyeci adı","nakliye"])
        nak_listesi = ["— Tümü —"]
        if nak_col_g:
            uniq = sorted(df_tum[nak_col_g].dropna().astype(str).str.strip().str.upper().unique().tolist())
            nak_listesi += [n for n in uniq if n and n != "NAN"]

        ana = self._ozel_frame
        ana.grid_rowconfigure(1, weight=1)
        ana.grid_columnconfigure(0, weight=1)

        # ── Filtre bandı ────────────────────────────────────────────────────
        filtre_kart = ctk.CTkFrame(ana, fg_color=C["card"], corner_radius=10,
                                   border_width=1, border_color=C["border"])
        filtre_kart.grid(row=0, column=0, sticky="ew", pady=(0,6))

        for i, lb in enumerate(["Nakliyeci", "Durum", "", "", ""]):
            ctk.CTkLabel(filtre_kart, text=lb, font=("Arial",10),
                         text_color=C["metin_dim"]
                         ).grid(row=0, column=i, padx=8, pady=(8,2), sticky="w")

        nak_var   = ctk.StringVar(value="— Tümü —")
        durum_var = ctk.StringVar(value="Tümü")

        nak_om = ctk.CTkOptionMenu(
            filtre_kart, variable=nak_var,
            values=nak_listesi,
            width=260, height=30, font=("Arial",10),
            fg_color=C["card2"], button_color=C["border"],
            dropdown_fg_color=C["card"],
            command=lambda v: _guncelle()
        )
        nak_om.grid(row=1, column=0, padx=8, pady=(0,10), sticky="ew")

        ctk.CTkOptionMenu(
            filtre_kart, variable=durum_var,
            values=["Tümü","Onay","Pregate Red","Sevkiyat Red","Operasyon Red","Red (Hepsi)"],
            width=170, height=30, font=("Arial",11),
            fg_color=C["card2"], button_color=C["border"],
            dropdown_fg_color=C["card"],
            command=lambda v: _guncelle()
        ).grid(row=1, column=1, padx=8, pady=(0,10), sticky="ew")

        ctk.CTkButton(
            filtre_kart, text="🔄  Sıfırla",
            fg_color=C["card2"], hover_color=C["border"],
            text_color=C["metin_dim"], font=("Arial",10), width=80, height=30,
            command=lambda: (nak_var.set("— Tümü —"),
                             durum_var.set("Tümü"), _guncelle())
        ).grid(row=1, column=2, padx=4, pady=(0,10))

        ctk.CTkButton(
            filtre_kart, text="📥  Excel",
            fg_color="#7D6608", hover_color="#5D4E08",
            text_color=C["altin"], font=("Arial",10), width=80, height=30,
            command=lambda: self._nakliye_excel(_df_hazirla(), None)
        ).grid(row=1, column=3, padx=4, pady=(0,10))

        lbl_sayac = ctk.CTkLabel(filtre_kart, text="", font=("Arial",9),
                                  text_color=C["metin_koyu"])
        lbl_sayac.grid(row=1, column=4, padx=8, pady=(0,10), sticky="w")

        # ── Tablo ────────────────────────────────────────────────────────────
        tablo_c = ctk.CTkFrame(ana, fg_color=C["card"], corner_radius=10)
        tablo_c.grid(row=1, column=0, sticky="nsew")
        tablo_c.grid_rowconfigure(0, weight=1)
        tablo_c.grid_columnconfigure(0, weight=1)

        sy = ttk.Scrollbar(tablo_c, orient="vertical")
        sy.grid(row=0, column=1, sticky="ns")
        sx = ttk.Scrollbar(tablo_c, orient="horizontal")
        sx.grid(row=1, column=0, sticky="ew")
        tablo = ttk.Treeview(tablo_c, yscrollcommand=sy.set,
                             xscrollcommand=sx.set, show="headings")
        tablo.grid(row=0, column=0, sticky="nsew", padx=2, pady=2)
        sy.configure(command=tablo.yview)
        sx.configure(command=tablo.xview)

        tablo.bind("<Double-1>", lambda e: _detay_popup())

        from tkinter import Menu as TkMenu
        ctx = TkMenu(self, tearoff=0, bg="#1A2438", fg=C["metin"],
                     activebackground=C["mavi"], activeforeground="#FFF",
                     font=("Arial",10))
        ctx.add_command(label="🔍  Nakliyeci Detay", command=lambda: _detay_popup())
        ctx.add_command(label="📋  Seçili Kopyala",
                        command=lambda: self._genel_tablo_kopyala(tablo, False))
        ctx.add_command(label="📄  Tümünü Kopyala",
                        command=lambda: self._genel_tablo_kopyala(tablo, True))

        def _sag_tik(e):
            item = tablo.identify_row(e.y)
            if item: tablo.selection_set(item)
            try: ctx.tk_popup(e.x_root, e.y_root)
            finally: ctx.grab_release()
        tablo.bind("<Button-3>", _sag_tik)
        tablo.bind("<Button-2>", _sag_tik)

        df_aktif_ref   = [pd.DataFrame()]
        goster_ref     = [[]]

        def _df_hazirla():
            df = df_tum.copy()
            # Nakliyeci filtresi
            nak = nak_var.get()
            if nak != "— Tümü —" and nak_col_g and nak_col_g in df.columns:
                df = df[df[nak_col_g].astype(str).str.strip().str.upper() == nak]
            # İptal Adımı filtresi
            durum = durum_var.get()
            adim_c = sutun_bul(df, ["iptal adimi", "iptal adımı", "İptal Adımı"])
            if durum != "Tümü" and adim_c:
                sa = df[adim_c].astype(str).str.strip()
                if durum == "Onay":
                    df = df[~sa.isin(TUM_RED_ADIMLARI)]
                elif durum == "Pregate Red":
                    df = df[sa.isin(PREGATE_RED_ADIMLARI)]
                elif durum == "Sevkiyat Red":
                    df = df[sa.isin(SEVKIYAT_RED_ADIMLARI)]
                elif durum == "Operasyon Red":
                    df = df[sa.isin(OPERASYON_RED_ADIMLARI)]
                elif durum == "Red (Hepsi)":
                    df = df[sa.isin(TUM_RED_ADIMLARI)]
            return df

        def _guncelle():
            df = _df_hazirla()
            df_aktif_ref[0] = df

            # Sütun sırası: Nakliyeci, İptal Açıklama, Tarih, Çekici, Durum, sonrası
            tum = [c for c in df.columns if not c.startswith("_")]
            onc = ["nakliyeci adi","nakliyeci adı","nakliye",
                   "iptal aciklama","aciklama","neden",
                   "tarih","date","kayıt",
                   "çekici","cekici","plaka","durum","departman"]
            g = []
            for anh in onc:
                c = sutun_bul(df, [anh])
                if c and c not in g: g.append(c)
            for c in tum[:10]:
                if c not in g: g.append(c)
            g = [c for c in g if c in df.columns] or tum[:6]
            goster_ref[0] = g

            tablo.delete(*tablo.get_children())
            tablo["columns"] = g
            gen = {"nakliyeci":200,"aciklama":230,"tarih":130,
                   "çekici":90,"plaka":90,"durum":120,"departman":130}
            for c in g:
                ck = c.lower().replace(" ","")
                w = next((v for k,v in gen.items() if k in ck), 110)
                tablo.column(c, anchor="w", width=w, minwidth=50)
                tablo.heading(c, text=c.replace("\n"," "), anchor="w")

            for i, (_, row) in enumerate(df.head(500).iterrows()):
                vals = []
                for c in g:
                    val = row.get(c,"")
                    try:
                        if pd.notna(val) and hasattr(val,"strftime"):
                            val = val.strftime("%d.%m.%Y %H:%M")
                    except Exception: pass
                    s = str(val)
                    vals.append(s[:80] if s not in ("nan","NaT","None") else "")
                tablo.insert("","end", values=vals,
                             tags=("tek_r" if i%2==0 else "cift_r",))

            if TEMA_MOD == "Light":
                tablo.tag_configure("tek_r",  background="#F7FAFC", foreground="#1A202C", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#EDF2F7", foreground="#1A202C", font=("Consolas",10))
            else:
                tablo.tag_configure("tek_r",  background="#141C2B", foreground="#E8EDF5", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#111827", foreground="#E8EDF5", font=("Consolas",10))

            lbl_sayac.configure(
                text=f"  {len(df)}/{len(df_tum)} kayıt  |  Çift tıkla → detay")

        def _detay_popup():
            sel = tablo.selection()
            if not sel:
                messagebox.showinfo("Bilgi","Önce bir satır seçin."); return
            idx = tablo.index(sel[0])
            df_a = df_aktif_ref[0]
            if df_a.empty or idx >= len(df_a): return
            row = df_a.iloc[idx]
            nak_c = sutun_bul(df_a, ["nakliyeci adi","nakliyeci adı","nakliye"])
            nak_adi = str(row[nak_c]).strip().upper() if nak_c else "?"
            df_nak = df_tum[df_tum[nak_c].astype(str).str.strip().str.upper() == nak_adi].copy() if nak_c else df_a.copy()

            pop = ctk.CTkToplevel(self)
            pop.title(f"🚛  {nak_adi}  —  {len(df_nak)} Kayıt")
            pop.geometry("1100x680")
            pop.resizable(True, True)
            pop.configure(fg_color=C["bg"])
            pop.grab_set(); pop.lift(); pop.focus_force()
            pop.grid_rowconfigure(2, weight=1)
            pop.grid_columnconfigure(0, weight=1)

            bb = ctk.CTkFrame(pop, fg_color=C["terminal"], corner_radius=0, height=44)
            bb.grid(row=0, column=0, sticky="ew"); bb.pack_propagate(False)
            ctk.CTkLabel(bb, text=f"  🚛  {nak_adi}  —  {len(df_nak)} Kayıt",
                         font=("Arial",13,"bold"), text_color="#FFFFFF").pack(side="left", padx=10)
            ctk.CTkButton(bb, text="✕  Kapat", fg_color="#0C3028",
                          text_color="#FFFFFF", font=("Arial",11), width=90, height=30,
                          command=pop.destroy).pack(side="right", padx=6, pady=7)
            ctk.CTkButton(bb, text="📥  Excel", fg_color="#7D6608",
                          text_color=C["altin"], font=("Arial",11), width=80, height=30,
                          command=lambda: self._firma_excele_kaydet(df_nak, nak_adi)
                          ).pack(side="right", padx=4, pady=7)

            ff = ctk.CTkFrame(pop, fg_color=C["card2"], corner_radius=0, height=44)
            ff.grid(row=1, column=0, sticky="ew"); ff.pack_propagate(False)
            d2_var = ctk.StringVar(value="Tümü")
            ctk.CTkLabel(ff, text="  Durum:", font=("Arial",10),
                         text_color=C["metin_dim"]).pack(side="left", padx=(8,2), pady=10)
            ctk.CTkOptionMenu(ff, variable=d2_var,
                              values=["Tümü","Onay","Pregate Red","Sevkiyat Red","Operasyon Red","Red (Hepsi)"],
                              width=170, height=28, font=("Arial",10),
                              fg_color=C["card"], button_color=C["border"],
                              dropdown_fg_color=C["card"],
                              command=lambda v: _pg()
                              ).pack(side="left", padx=4, pady=10)
            lbl_p = ctk.CTkLabel(ff, text="", font=("Arial",9), text_color=C["metin_koyu"])
            lbl_p.pack(side="left", padx=10)

            tc = ctk.CTkFrame(pop, fg_color=C["card"], corner_radius=0)
            tc.grid(row=2, column=0, sticky="nsew")
            tc.grid_rowconfigure(0, weight=1); tc.grid_columnconfigure(0, weight=1)
            sy2 = ttk.Scrollbar(tc, orient="vertical"); sy2.grid(row=0, column=1, sticky="ns")
            sx2 = ttk.Scrollbar(tc, orient="horizontal"); sx2.grid(row=1, column=0, sticky="ew")
            t2 = ttk.Treeview(tc, yscrollcommand=sy2.set, xscrollcommand=sx2.set, show="headings")
            t2.grid(row=0, column=0, sticky="nsew")
            sy2.configure(command=t2.yview); sx2.configure(command=t2.xview)

            def _pg():
                d = df_nak.copy()
                df_fil = d2_var.get()
                adim_c2 = sutun_bul(d, ["iptal adimi", "iptal adımı", "İptal Adımı"])
                if df_fil != "Tümü" and adim_c2:
                    sa2 = d[adim_c2].astype(str).str.strip()
                    if df_fil == "Onay":
                        d = d[~sa2.isin(TUM_RED_ADIMLARI)]
                    elif df_fil == "Pregate Red":
                        d = d[sa2.isin(PREGATE_RED_ADIMLARI)]
                    elif df_fil == "Sevkiyat Red":
                        d = d[sa2.isin(SEVKIYAT_RED_ADIMLARI)]
                    elif df_fil == "Operasyon Red":
                        d = d[sa2.isin(OPERASYON_RED_ADIMLARI)]
                    elif df_fil == "Red (Hepsi)":
                        d = d[sa2.isin(TUM_RED_ADIMLARI)]
                tum2 = [c for c in d.columns if not c.startswith("_")]
                onc2 = ["nakliyeci adi","nakliyeci adı","nakliye",
                        "iptal aciklama","aciklama","neden",
                        "tarih","date","kayıt","çekici","cekici","plaka","durum","departman"]
                g2 = []
                for anh in onc2:
                    c = sutun_bul(d, [anh])
                    if c and c not in g2: g2.append(c)
                for c in tum2[:10]:
                    if c not in g2: g2.append(c)
                g2 = [c for c in g2 if c in d.columns] or tum2[:6]
                t2.delete(*t2.get_children())
                t2["columns"] = g2
                gen2 = {"nakliyeci":200,"aciklama":230,"tarih":130,"çekici":90,"plaka":90,"durum":120,"departman":130}
                for c in g2:
                    ck = c.lower().replace(" ","")
                    w = next((v for k,v in gen2.items() if k in ck), 110)
                    t2.column(c, anchor="w", width=w, minwidth=50)
                    t2.heading(c, text=c.replace("\n"," "), anchor="w")
                for i, (_, rw) in enumerate(d.head(500).iterrows()):
                    vals = []
                    for c in g2:
                        val = rw.get(c,"")
                        try:
                            if pd.notna(val) and hasattr(val,"strftime"):
                                val = val.strftime("%d.%m.%Y %H:%M")
                        except Exception: pass
                        s = str(val)
                        vals.append(s[:80] if s not in ("nan","NaT","None") else "")
                    t2.insert("","end", values=vals, tags=("tek_r" if i%2==0 else "cift_r",))
                if TEMA_MOD == "Light":
                    t2.tag_configure("tek_r",  background="#F7FAFC", foreground="#1A202C", font=("Consolas",10))
                    t2.tag_configure("cift_r", background="#EDF2F7", foreground="#1A202C", font=("Consolas",10))
                else:
                    t2.tag_configure("tek_r",  background="#141C2B", foreground="#E8EDF5", font=("Consolas",10))
                    t2.tag_configure("cift_r", background="#111827", foreground="#E8EDF5", font=("Consolas",10))
                lbl_p.configure(text=f"  {len(d)}/{len(df_nak)} kayıt")
            _pg()

        _guncelle()

    def _nakliye_excel(self, df, secili_sutunlar):
        """Nakliyeci tablosunu Excel olarak kaydeder."""
        if df is None or df.empty:
            messagebox.showwarning("Uyarı","Veri yok.")
            return
        goster = secili_sutunlar or [c for c in df.columns if not c.startswith("_")]
        goster = [c for c in goster if c in df.columns]
        yol = filedialog.asksaveasfilename(
            title="Nakliyeci Listesini Kaydet",
            defaultextension=".xlsx",
            initialfile=f"Nakliyeci_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel","*.xlsx")]
        )
        if not yol: return
        try:
            df[goster].to_excel(yol, index=False)
            self._bildirim(f"✔  Nakliyeci listesi kaydedildi.")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")

    def _nakliye_tum(self):
        """Eski metod — yeni _goster_nakliyeci tarafından karşılanıyor."""
        self._goster_nakliyeci()

    def _nakliye_onay(self):
        """Eski metod — yeni sistem üzerinden filtrele."""
        self._goster_nakliyeci()

    def _nakliye_red(self):
        """Nakliyeci Red Matrix — nakliyeci × iptal nedeni pivot, popup olarak."""
        if self.motor.df_pregate.empty:
            messagebox.showwarning("Uyarı", "Pregate dosyası yüklü değil."); return

        pivot = self.motor.nakliyeci_matrix()
        if pivot.empty:
            messagebox.showwarning("Uyarı",
                "Matrix için yeterli veri bulunamadı.\n"
                "Durum / Nakliyeci / Açıklama sütunlarının varlığını kontrol edin.")
            return

        pop = ctk.CTkToplevel(self)
        pop.title("📊  Nakliyeci Red Matrix  —  Kategori Pivot Dağılımı")
        pop.geometry("1300x720")
        pop.resizable(True, True)
        pop.configure(fg_color=C["bg"])
        pop.grab_set(); pop.lift(); pop.focus_force()

        pop.grid_rowconfigure(1, weight=1)
        pop.grid_columnconfigure(0, weight=1)

        # Başlık
        bb = ctk.CTkFrame(pop, fg_color="#3D100C", corner_radius=0, height=44)
        bb.grid(row=0, column=0, sticky="ew"); bb.pack_propagate(False)
        ctk.CTkLabel(bb, text=f"  📊  Nakliyeci Red Matrix  —  {len(pivot)-1} Firma",
                     font=("Arial",13,"bold"), text_color=C["altin"]).pack(side="left", padx=10)
        ctk.CTkButton(bb, text="✕  Kapat", fg_color="#500A0A", hover_color="#3A0606",
                      text_color="#FFFFFF", font=("Arial",11), width=90, height=30,
                      command=pop.destroy).pack(side="right", padx=6, pady=7)
        ctk.CTkButton(bb, text="📥  Excel İndir", fg_color="#7D6608", hover_color="#5D4E08",
                      text_color=C["altin"], font=("Arial",11), width=110, height=30,
                      command=lambda: self._matrix_excel_indir(pivot)
                      ).pack(side="right", padx=4, pady=7)
        ctk.CTkButton(bb, text="📋  Kopyala", fg_color="#1A3A5C", hover_color="#0F2540",
                      text_color="#FFFFFF", font=("Arial",11), width=90, height=30,
                      command=lambda: self._genel_tablo_kopyala(tablo, True)
                      ).pack(side="right", padx=4, pady=7)

        # Filtre — dropdown
        filtre_f = ctk.CTkFrame(pop, fg_color=C["card2"], corner_radius=0, height=46)
        filtre_f.grid(row=1, column=0, sticky="ew"); filtre_f.pack_propagate(False)

        # Nakliyeci listesi pivot index'ten
        firma_listesi = ["— Tümü —"] + [str(idx) for idx in pivot.index if idx != "TOPLAM"]

        nak_var = ctk.StringVar(value="— Tümü —")
        ctk.CTkLabel(filtre_f, text="  Nakliyeci:", font=("Arial",10),
                     text_color=C["metin_dim"]).pack(side="left", padx=(8,2), pady=10)
        nak_om = ctk.CTkOptionMenu(
            filtre_f, variable=nak_var,
            values=firma_listesi,
            width=280, height=28, font=("Arial",10),
            fg_color=C["card"], button_color=C["border"],
            dropdown_fg_color=C["card"],
            command=lambda v: _filtrele()
        )
        nak_om.pack(side="left", padx=4, pady=10)
        ctk.CTkButton(filtre_f, text="✕  Sıfırla", fg_color=C["card2"],
                      hover_color=C["border"], text_color=C["metin_dim"],
                      font=("Arial",10), width=80, height=28,
                      command=lambda: (nak_var.set("— Tümü —"), _filtrele())
                      ).pack(side="left", padx=4)
        lbl_sayac = ctk.CTkLabel(filtre_f, text="", font=("Arial",9),
                                  text_color=C["metin_koyu"])
        lbl_sayac.pack(side="left", padx=10)

        # Tablo
        tablo_c = ctk.CTkFrame(pop, fg_color=C["card"], corner_radius=0)
        tablo_c.grid(row=2, column=0, sticky="nsew")
        pop.grid_rowconfigure(2, weight=1)
        tablo_c.grid_rowconfigure(0, weight=1); tablo_c.grid_columnconfigure(0, weight=1)
        sy = ttk.Scrollbar(tablo_c, orient="vertical"); sy.grid(row=0, column=1, sticky="ns")
        sx = ttk.Scrollbar(tablo_c, orient="horizontal"); sx.grid(row=1, column=0, sticky="ew")
        tablo = ttk.Treeview(tablo_c, yscrollcommand=sy.set, xscrollcommand=sx.set, show="headings")
        tablo.grid(row=0, column=0, sticky="nsew")
        sy.configure(command=tablo.yview); sx.configure(command=tablo.xview)

        from tkinter import Menu as TkM
        m = TkM(pop, tearoff=0, bg="#1A2438", fg="white",
                activebackground=C["mavi"], activeforeground="white", font=("Arial",10))
        m.add_command(label="📋  Seçili Kopyala", command=lambda: self._genel_tablo_kopyala(tablo, False))
        m.add_command(label="📄  Tümünü Kopyala", command=lambda: self._genel_tablo_kopyala(tablo, True))
        m.add_command(label="📥  Excel", command=lambda: self._matrix_excel_indir(pivot))
        def _stk(e):
            item = tablo.identify_row(e.y)
            if item: tablo.selection_set(item)
            try: m.tk_popup(e.x_root, e.y_root)
            finally: m.grab_release()
        tablo.bind("<Button-3>", _stk); tablo.bind("<Button-2>", _stk)

        def _filtrele():
            secili = nak_var.get()
            # Dropdown seçimine göre filtrele
            if secili == "— Tümü —":
                goster_idx = list(pivot.index)
            else:
                goster_idx = [secili, "TOPLAM"] if "TOPLAM" in pivot.index else [secili]

            cols = ["Nakliyeci Firma"] + list(pivot.columns)
            tablo["columns"] = cols
            tablo.column("Nakliyeci Firma", anchor="w", width=240, minwidth=120)
            tablo.heading("Nakliyeci Firma", text="Nakliyeci Firma", anchor="w")
            for c in pivot.columns:
                if c in ("Toplam Red", "Toplam Giriş"):
                    w = 110
                elif c in ("PREGATE RED", "SEVKİYAT RED", "OPERASYON RED"):
                    w = 110
                else:
                    w = 130
                tablo.column(c, anchor="center", width=w, minwidth=60)
                tablo.heading(c, text=c[:20], anchor="center")

            tablo.delete(*tablo.get_children())
            firma_say = 0
            for i, idx in enumerate(goster_idx):
                row = pivot.loc[idx]
                vals = [idx] + [int(row[c]) if row[c] != 0 else "" for c in pivot.columns]
                if idx == "TOPLAM":
                    tag = "toplam"
                else:
                    tag = "tek_r" if i%2==0 else "cift_r"
                    firma_say += 1
                tablo.insert("","end", values=vals, tags=(tag,))

            if TEMA_MOD == "Light":
                tablo.tag_configure("tek_r",  background="#F7FAFC", foreground="#1A202C", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#EDF2F7", foreground="#1A202C", font=("Consolas",10))
            else:
                tablo.tag_configure("tek_r",  background="#141C2B", foreground="#E8EDF5", font=("Consolas",10))
                tablo.tag_configure("cift_r", background="#111827", foreground="#E8EDF5", font=("Consolas",10))
            tablo.tag_configure("toplam", background="#0A0E17",
                                foreground=C["altin"], font=("Arial",10,"bold"))
            lbl_sayac.configure(text=f"  {firma_say} firma  |  {len(pivot.columns)} sütun")

        _filtrele()

    def _matrix_excel_indir(self, pivot):
        """Pivot matrix'i Excel olarak indir."""
        yol = filedialog.asksaveasfilename(
            title="Red Matrix Excel Kaydet",
            defaultextension=".xlsx",
            initialfile=f"Red_Matrix_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel","*.xlsx")])
        if not yol: return
        try:
            pivot.reset_index(names="Nakliyeci Firma").to_excel(yol, index=False)
            messagebox.showinfo("Başarılı", f"Kaydedildi:\n{yol}")
        except Exception as e:
            messagebox.showerror("Hata", f"Kaydedilemedi:\n{e}")

    # =========================================================================
    # EKSİK VERİ UYARISI (metin sayfası alt kısmı)
    # =========================================================================
    def _eksik_uyari_ekle(self, df):
        eksikler = self.motor.eksik_analiz(
            df, ["nakliye","durum","aciklama","havuz"])
        if eksikler:
            self._yaz("\n  ⚠  VERİ KALİTE UYARILARI  \n", "h1")
            for col, sayi in eksikler.items():
                self._yaz(
                    f"  ⚠  '{col}'  sütununda  {sayi}  boş / eksik satır bulundu.\n",
                    "uyari")

    # =========================================================================
    # EXCEL EXPORT  —  Kapsamlı çok sayfalı rapor
    # =========================================================================
    def _export_excel(self):
        herhangi = any([
            not self.motor.df_dinamik.empty,
            not self.motor.df_pregate.empty,
            not self.motor.df_ek.empty,
        ])
        if not herhangi:
            messagebox.showwarning("Uyarı", "Hiçbir veri dosyası yüklü değil.")
            return

        yol = filedialog.asksaveasfilename(
            title="Excel Raporunu Kaydet",
            defaultextension=".xlsx",
            initialfile=f"SeymenRaporlama_{datetime.datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )
        if not yol:
            return

        try:
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                         Border, Side)
            from openpyxl.utils import get_column_letter

            wb_writer = pd.ExcelWriter(yol, engine="openpyxl")

            # ─── Renk sabitleri ───────────────────────────────────────────────
            KOYU_BG   = "0A0E17"
            ALTIN     = "D4AF37"
            MAVI      = "2E86DE"
            YESIL     = "26A65B"
            KIRMIZI   = "C0392B"
            PRE_BG2   = "1A0A0A"      # Pregate Red sayfası başlık bg
            SEV_BG    = "0A0A1A"      # Sevkiyat Red
            TER_BG    = "0A1A1A"      # Terminal Red
            PRG_BG    = "1A150A"      # Pregate Red
            SATIR1    = "141C2B"
            SATIR2    = "111827"
            BASLIK_BG = "0D1117"

            def _stil_uygula(ws, df_veri, baslik_renk=MAVI,
                             satir1=SATIR1, satir2=SATIR2):
                """Sayfaya koyu tema stili uygular."""
                # Başlık satırı
                for col_idx, col_ad in enumerate(df_veri.columns, 1):
                    hucre = ws.cell(row=1, column=col_idx)
                    hucre.fill = PatternFill("solid", fgColor=BASLIK_BG)
                    hucre.font = Font(color=baslik_renk, bold=True,
                                     size=11, name="Calibri")
                    hucre.alignment = Alignment(horizontal="center",
                                                vertical="center",
                                                wrap_text=True)
                    hucre.border = Border(
                        bottom=Side(style="thin", color="252D3D"))
                    ws.row_dimensions[1].height = 22

                # Veri satırları
                for row_idx in range(2, ws.max_row + 1):
                    bg = satir1 if row_idx % 2 == 0 else satir2
                    for col_idx in range(1, ws.max_column + 1):
                        h = ws.cell(row=row_idx, column=col_idx)
                        h.fill = PatternFill("solid", fgColor=bg)
                        h.font = Font(color="E8EDF5", size=10, name="Calibri")
                        h.alignment = Alignment(horizontal="center",
                                                vertical="center")
                    ws.row_dimensions[row_idx].height = 18

                # Sütun genişlikleri
                for col_idx, col_ad in enumerate(df_veri.columns, 1):
                    harf = get_column_letter(col_idx)
                    max_w = max(
                        len(str(col_ad)),
                        df_veri[col_ad].astype(str).str.len().max()
                        if len(df_veri) > 0 else 0
                    )
                    ws.column_dimensions[harf].width = min(max_w + 4, 40)

                # Sayfa arka planı (tab rengi)
                ws.sheet_properties.tabColor = baslik_renk[:6]

            def _yaz(df_veri, sayfa_adi, baslik_renk=MAVI):
                if df_veri is None or df_veri.empty:
                    return
                df_veri.to_excel(wb_writer, sheet_name=sayfa_adi,
                                 index=False)
                ws = wb_writer.sheets[sayfa_adi]
                _stil_uygula(ws, df_veri, baslik_renk)

            # ─── 1. ÖZET SAYIM ────────────────────────────────────────────────
            ozet_df = self.motor.ozet_sayim_tablosu()
            _yaz(ozet_df, "📊 Genel Özet", ALTIN)

            # ─── 2. NAKLİYECİ RED MATRİX AÇIKLAMA SAYFASI ───────────────────
            aciklama_data = {
                "Sütun Adı": [
                    "EVRAKSAL",
                    "FİRMA TALEBİ İPTAL",
                    "SÜRÜCÜ",
                    "DONANIMSAL",
                    "KKD",
                    "DİĞER",
                    "",
                    "PREGATE RED",
                    "SEVKİYAT RED",
                    "OPERASYON RED",
                    "",
                    "Toplam Red",
                    "Toplam Giriş",
                    "",
                    "ÖNEMLİ NOT",
                ],
                "Kaynak / Açıklama": [
                    "Ana Neden = EVRAKSAL olan red kayıtları (eksik/geçersiz evrak)",
                    "Ana Neden = FİRMA TALEBİ İPTAL olan kayıtlar (firma kendi isteğiyle iptal)",
                    "Ana Neden = SÜRÜCÜ olan kayıtlar (sürücü kaynaklı red)",
                    "Ana Neden = DONANIMSAL olan kayıtlar (araç/donanım eksikliği)",
                    "Ana Neden = KKD olan kayıtlar (Kişisel Koruyucu Donanım eksikliği)",
                    "Ana Neden listesinde yer almayan tüm diğer red kayıtları",
                    "",
                    "İptal Adımı = 'PregateRed' VEYA 'PregateKayitRed' olan satırlar",
                    "İptal Adımı = 'SevkiyatPlanlama' olan satırlar",
                    "İptal Adımı = 'PlatformRed' VEYA 'TerminalRed' olan satırlar",
                    "",
                    "= PREGATE RED + SEVKİYAT RED + OPERASYON RED toplamı",
                    "= Onaylanan + Reddedilen tüm araç girişleri (o nakliyeciye ait)",
                    "",
                    "Ana Neden sütunlarının toplamı her zaman Toplam Red ile eşit olmalıdır. "                    "Eşit değilse veri dosyasındaki Ana Neden sütununda "                    "tanımsız değerler vardır — bunlar DİĞER sütununa aktarılmıştır.",
                ],
                "Kaynak Sütun (Excel'de)": [
                    "Ana Neden",
                    "Ana Neden",
                    "Ana Neden",
                    "Ana Neden",
                    "Ana Neden",
                    "Ana Neden",
                    "",
                    "İptal Adımı",
                    "İptal Adımı",
                    "İptal Adımı",
                    "",
                    "Hesaplanmış",
                    "Tüm kayıtlar",
                    "",
                    "",
                ],
            }
            aciklama_df = pd.DataFrame(aciklama_data)
            aciklama_df.to_excel(wb_writer, sheet_name="ℹ Açıklama", index=False)
            ws_ac = wb_writer.sheets["ℹ Açıklama"]
            # Başlık stili
            for ci, col in enumerate(aciklama_df.columns, 1):
                h = ws_ac.cell(row=1, column=ci)
                h.fill = PatternFill("solid", fgColor="0D1117")
                h.font = Font(color=ALTIN, bold=True, size=11, name="Calibri")
                h.alignment = Alignment(horizontal="center", vertical="center")
                h.border = Border(bottom=Side(style="thin", color="252D3D"))
            ws_ac.row_dimensions[1].height = 22
            # Veri satırları
            for ri in range(2, ws_ac.max_row + 1):
                val0 = str(ws_ac.cell(row=ri, column=1).value or "")
                if val0 == "":
                    bg = "080C14"
                elif val0 in ("PREGATE RED", "SEVKİYAT RED", "OPERASYON RED",
                               "Toplam Red", "Toplam Giriş"):
                    bg = "0A1520"
                elif val0 == "ÖNEMLİ NOT":
                    bg = "1A0A00"
                else:
                    bg = "141C2B" if ri % 2 == 0 else "111827"
                for ci in range(1, ws_ac.max_column + 1):
                    h = ws_ac.cell(row=ri, column=ci)
                    h.fill = PatternFill("solid", fgColor=bg)
                    font_col = "E8EDF5"
                    if val0 in ("PREGATE RED", "SEVKİYAT RED", "OPERASYON RED"):
                        font_col = "E87070"
                    elif val0 == "Toplam Red":
                        font_col = "E8A070"
                    elif val0 == "ÖNEMLİ NOT":
                        font_col = ALTIN
                    h.font = Font(color=font_col, size=10, name="Calibri",
                                  bold=(val0 == "ÖNEMLİ NOT"))
                    h.alignment = Alignment(horizontal="left", vertical="center",
                                            wrap_text=True)
                ws_ac.row_dimensions[ri].height = 36 if val0 == "ÖNEMLİ NOT" else 20
            # Sütun genişlikleri
            ws_ac.column_dimensions["A"].width = 28
            ws_ac.column_dimensions["B"].width = 75
            ws_ac.column_dimensions["C"].width = 22
            ws_ac.sheet_properties.tabColor = ALTIN[:6]

            # ─── 3. DİNAMİK HAM VERİ ─────────────────────────────────────────
            if not self.motor.df_dinamik.empty:
                _yaz(self.motor.df_dinamik, "Dinamik Ham Veri", MAVI)

            # ─── 3-5. DİNAMİK DEPARTMAN ÖZETLERİ ────────────────────────────
            for dep, renk in [("Kimya","7B241C"),
                               ("Dow","154360"),
                               ("Diğer","5D4037")]:
                df_d = self.motor._dinamik_filtrele(dep)
                if not df_d.empty:
                    _yaz(df_d, f"Dinamik {dep}", renk)

            # ─── 6. PREGATE HAM VERİ ─────────────────────────────────────────
            if not self.motor.df_pregate.empty:
                _yaz(self.motor.df_pregate, "Pregate Ham Veri", "2E86DE")

            # ─── 7. TERMİNAL HAM VERİ ────────────────────────────────────────
            term_df = self.motor._pregate_filtrele("Poliport Terminal")
            if not term_df.empty:
                _yaz(term_df, "Terminal Ham Veri", "0E6251")

            # ─── 8-10. RED TİPİ BAZLI NAKLİYECİ LİSTELERİ ───────────────────
            # Nakliyeci çalışma kopyası hazırla (terminal)
            if not self.motor.df_pregate.empty:
                dep_col = sutun_bul(self.motor.df_pregate, ["departman"])
                if dep_col:
                    sd = self.motor.df_pregate[dep_col].astype(str).str.lower()
                    self.motor.df_pol_tmp = self.motor.df_pregate[
                        sd.str.contains("terminal", na=False)].copy()
                else:
                    self.motor.df_pol_tmp = self.motor.df_pregate.copy()

            red_listeler = self.motor.red_tipi_nakliyeci_listesi("Poliport Terminal")

            if red_listeler:
                # Pregate Red
                prg = red_listeler.get("pregate_red", pd.DataFrame())
                if not prg.empty:
                    _yaz(prg, "🔴 Pregate Red", KIRMIZI)

                # Sevkiyat Red
                sev = red_listeler.get("sevkiyat_red", pd.DataFrame())
                if not sev.empty:
                    _yaz(sev, "🟠 Sevkiyat Red", "E67E22")

                # Operasyon Red
                op = red_listeler.get("operasyon_red", pd.DataFrame())
                if not op.empty:
                    _yaz(op, "🔵 Operasyon Red", "2E86DE")

                # Genel nakliyeci red matrix
                mx = red_listeler.get("genel_matrix", pd.DataFrame())
                if not mx.empty:
                    mx_reset = mx.reset_index()
                    _yaz(mx_reset, "🗂 Nakliyeci Red Matrix", "7B241C")

                # ── Detay Red sayfaları (İptal Açıklama bazlı) ───────────────
                detay = self.motor.detay_red_matrix("Poliport Terminal")
                if detay:
                    for tip_key, sayfa_adi, renk in [
                        ("pregate",   "📋 Detay Pregate Red",   KIRMIZI),
                        ("sevkiyat",  "📋 Detay Sevkiyat Red",  "E67E22"),
                        ("operasyon", "📋 Detay Operasyon Red", "2E86DE"),
                    ]:
                        df_det = detay.get(tip_key, pd.DataFrame())
                        if df_det.empty:
                            continue
                        df_det_r = df_det.reset_index()
                        df_det_r.to_excel(wb_writer, sheet_name=sayfa_adi, index=False)
                        ws_d = wb_writer.sheets[sayfa_adi]
                        # Başlık satırı
                        for ci in range(1, ws_d.max_column + 1):
                            h = ws_d.cell(row=1, column=ci)
                            h.fill = PatternFill("solid", fgColor=BASLIK_BG)
                            h.font = Font(color=renk if isinstance(renk, str) else KIRMIZI,
                                          bold=True, size=10, name="Calibri")
                            h.alignment = Alignment(horizontal="center",
                                                    vertical="center", wrap_text=True)
                            h.border = Border(bottom=Side(style="thin", color="252D3D"))
                        ws_d.row_dimensions[1].height = 28
                        # Veri satırları
                        for ri in range(2, ws_d.max_row + 1):
                            val0 = str(ws_d.cell(row=ri, column=1).value or "")
                            bg = "1A0A0A" if val0 == "TOPLAM" else (
                                 SATIR1 if ri % 2 == 0 else SATIR2)
                            fnt_bold = (val0 == "TOPLAM")
                            for ci in range(1, ws_d.max_column + 1):
                                h = ws_d.cell(row=ri, column=ci)
                                h.fill = PatternFill("solid", fgColor=bg)
                                h.font = Font(color="E8EDF5", size=10,
                                              name="Calibri", bold=fnt_bold)
                                h.alignment = Alignment(horizontal="center",
                                                        vertical="center")
                            ws_d.row_dimensions[ri].height = 18
                        # Sütun genişlikleri — 1. sütun (Nakliyeci) geniş,
                        # detay sütunları orta, son sütun (Toplam) dar
                        ws_d.column_dimensions[get_column_letter(1)].width = 30
                        for ci in range(2, ws_d.max_column):
                            col_hdr = str(ws_d.cell(row=1, column=ci).value or "")
                            ws_d.column_dimensions[get_column_letter(ci)].width = min(
                                max(len(col_hdr) + 2, 14), 45)
                        ws_d.column_dimensions[
                            get_column_letter(ws_d.max_column)].width = 10
                        ws_d.sheet_properties.tabColor = renk[:6]

            # ─── 11. NAKLİYECİ ONAY LİSTESİ ──────────────────────────────────
            if not self.motor.df_pol_tmp.empty:
                durum_col = sutun_bul(self.motor.df_pol_tmp, ["durum"])
                if durum_col:
                    onay_df = self.motor.df_pol_tmp[
                        self.motor.df_pol_tmp[durum_col].astype(str
                        ).str.lower().str.strip() == "onay"
                    ].copy()
                    if not onay_df.empty:
                        _yaz(onay_df, "✅ Onaylanan Nakliyeciler", YESIL)

            # ─── 12. EK DOSYA ─────────────────────────────────────────────────
            if not self.motor.df_ek.empty:
                _yaz(self.motor.df_ek, "📄 Ek Dosya", "6E2F1A")

            wb_writer.close()

            # Sayfa sıralaması: Özet başa
            try:
                from openpyxl import load_workbook
                wb = load_workbook(yol)
                sayfa_sirasi = ["📊 Genel Özet", "ℹ Açıklama"]
                diger = [s for s in wb.sheetnames if s not in sayfa_sirasi]
                wb._sheets = [wb[s] for s in sayfa_sirasi
                              if s in wb.sheetnames] + \
                             [wb[s] for s in diger]
                wb.save(yol)
            except Exception:
                pass  # Sıralama başarısız olsa da dosya kaydedilmiş

            messagebox.showinfo(
                "Excel Raporu Hazır",
                f"Dosya başarıyla kaydedildi.\n\n"
                f"{os.path.basename(yol)}\n\n"
                f"İçerik:\n"
                f"  • Genel Özet\n"
                f"  • Açıklama (Matrix sütun tanımları)\n"
                f"  • Dinamik Ham Veri + Departman sayfaları\n"
                f"  • Pregate Ham Veri + Terminal sayfası\n"
                f"  • Pregate Red / Sevkiyat Red / Operasyon Red\n"
                f"  • Nakliyeci Red Matrix\n"
                f"  • Onaylanan Nakliyeciler\n"
                f"  • Ek Dosya (yüklüyse)")

        except Exception as e:
            messagebox.showerror("Hata", f"Excel oluşturulamadı:\n{e}")



# ============================================================================
# BAŞLATICI
# ============================================================================
if __name__ == "__main__":
    app = SeymenRaporlama()
    app.mainloop()